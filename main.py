import json
import threading
import time
import subprocess
import os
import re
import uuid
import requests
import cv2
import glob
import urllib.parse
import hashlib
import pickle
from datetime import datetime
from flask import Flask, render_template, jsonify, request, send_from_directory, send_file, session, redirect, url_for, make_response
from PIL import Image
from io import BytesIO
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

app = Flask(__name__)
app.secret_key = 'cqubepro_secret_key_2024'  # Change this in production

# Global variable to store camera data
cameras_data = []

# Global configuration - centralized settings
# This is the single source of truth for all URLs and server settings
# To change the camera base URL, update the "camera_base_url" value below
# To change the API base URL, update the "api_base_url" value below
GLOBAL_CONFIG = {
    "camera_base_url": "http://192.168.1.3:9000",  # Base URL for camera operations
    "api_base_url": "http://192.168.1.3:9000",    # Base URL for API operations
    "server": {
        "host": "0.0.0.0",
        "port": 5000,
        "debug": False
    }
}

# Usage examples:
# - Get camera base URL: get_camera_base_url()
# - Get API base URL: get_api_base_url()
# - Access server config: GLOBAL_CONFIG["server"]

# Global variable to store external API configuration (deprecated - use GLOBAL_CONFIG)
external_api_config = {
    "base_url": GLOBAL_CONFIG["api_base_url"]
}

# Global variable to store server configuration (deprecated - use GLOBAL_CONFIG)
server_config = GLOBAL_CONFIG["server"]

# MediaMTX process management
mediamtx_process = None
mediamtx_exe_path = os.path.abspath("mediamtx/mediamtx.exe")
mediamtx_config_file = os.path.abspath("mediamtx/mediamtx/mediamtx.yml")
# Global variable to store base video path
base_video_path = ""
# Dictionary to store active threads for each camera
active_threads = {}
# Dictionary to store stop flags for each camera
stop_flags = {}
# Global variable to store groups
groups_data = []

# Global variables for video processing
merge_queue = None
merge_worker_thread = None
merging_tasks = {}
video_config = {
    'upload_to_api': True,  # Enabled - API server is running
    'store_locally': True,
    'api_url': f"{GLOBAL_CONFIG['api_base_url']}/api/video/upload",  # Uses global config
    'api_key': 'your_api_key_here',
    'upload_retry_attempts': 3,
    'upload_timeout': 30,
    'upload_delay_between_files': 0.5,  # seconds
    'strict_validation': False,  # Disable strict validation for segment files
    'min_file_age_seconds': 15,  # Increase from 5 to 15 seconds
    'file_stability_check_seconds': 3  # Increase from 1 to 3 seconds
}
def get_camera_base_url():
    """Get the camera base URL from global configuration"""
    return GLOBAL_CONFIG['camera_base_url']

def get_api_base_url():
    """Get the API base URL from global configuration"""
    return GLOBAL_CONFIG['api_base_url']

def get_default_organization_id():
    """Get the default organization ID from config.json"""
    try:
        with open('config.json', 'r') as f:
            config_data = json.load(f)
        if 'organizations' in config_data and len(config_data['organizations']) > 0:
            return config_data['organizations'][0]['id']
    except Exception as e:
        print(f"⚠️ Error loading organization ID from config.json: {e}")
    return 18  # Default fallback

def fetch_cameras_from_api(organization_id=None):
    """Fetch cameras list from external API and update config.json"""
    try:
        if organization_id is None:
            organization_id = get_default_organization_id()
        
        # API endpoint for fetching cameras
        api_url = f"http://192.168.1.6:9000/cameras/?organization_id={organization_id}"
        
        print(f"Fetching cameras from API: {api_url}")
        
        # Make API request
        response = requests.get(api_url, timeout=10)
        
        if response.status_code == 200:
            cameras_data = response.json()
            print(f"Successfully fetched {len(cameras_data)} cameras from API")
            
            # Load current config
            with open('config.json', 'r') as f:
                config_data = json.load(f)
            
            # Update cameras in config
            config_data['cameras'] = cameras_data
            
            # Save updated config
            with open('config.json', 'w') as f:
                json.dump(config_data, f, indent=2)
            
            print(f"Updated config.json with {len(cameras_data)} cameras")
            return cameras_data
            
        else:
            print(f"API request failed with status {response.status_code}: {response.text}")
            return None
            
    except Exception as e:
        print(f"Error fetching cameras from API: {e}")
        return None

api_organization_id = get_default_organization_id()
print(f"Default Organization ID loaded from config: {api_organization_id}")

# Video status tracking
VIDEO_STATUS_FILE = "video_upload_status.pkl"
video_status = {}  # Dictionary to track video upload status

def load_video_status():
    """Load video upload status from file"""
    global video_status
    try:
        if os.path.exists(VIDEO_STATUS_FILE):
            with open(VIDEO_STATUS_FILE, 'rb') as f:
                video_status = pickle.load(f)
            print(f"📊 Loaded {len(video_status)} video status records")
        else:
            video_status = {}
            print("📊 No existing video status file found, starting fresh")
    except Exception as e:
        print(f"⚠️ Error loading video status: {e}")
        video_status = {}

def save_video_status():
    """Save video upload status to file"""
    try:
        with open(VIDEO_STATUS_FILE, 'wb') as f:
            pickle.dump(video_status, f)
    except Exception as e:
        print(f"⚠️ Error saving video status: {e}")

def update_video_status(file_path, status, error_msg=None, retry_count=None):
    """Update video upload status"""
    file_key = os.path.basename(file_path)
    current_retry = video_status.get(file_key, {}).get('retry_count', 0)
    video_status[file_key] = {
        'file_path': file_path,
        'status': status,  # 'pending', 'uploading', 'completed', 'failed'
        'timestamp': datetime.now().isoformat(),
        'error_msg': error_msg,
        'retry_count': retry_count if retry_count is not None else current_retry
    }
    save_video_status()

def process_pending_videos():
    """Process videos that were not successfully uploaded on startup"""
    print("🔄 Checking for pending video uploads...")
    
    pending_count = 0
    failed_count = 0
    
    # Create a list of keys to remove to avoid modifying dict during iteration
    keys_to_remove = []
    
    for file_key, status_info in video_status.items():
        if status_info['status'] in ['pending', 'failed']:
            file_path = status_info['file_path']
            
            # Check if file still exists
            if not os.path.exists(file_path):
                print(f"⚠️ File no longer exists, removing from status: {file_key}")
                keys_to_remove.append(file_key)
                continue
            
            # Check retry count (max 5 retries)
            if status_info['retry_count'] >= 5:
                print(f"❌ Max retries reached for {file_key}, marking as permanently failed")
                update_video_status(file_path, 'failed', 'Max retries exceeded')
                failed_count += 1
                continue
            
            print(f"🔄 Retrying upload for {file_key} (attempt {status_info['retry_count'] + 1})")
            pending_count += 1
            
            # Add to upload queue
            upload_queue.append(file_path)
    
    # Remove keys that were marked for deletion
    for key in keys_to_remove:
        if key in video_status:
            del video_status[key]
    
    if pending_count > 0:
        print(f"📤 Queued {pending_count} pending videos for upload")
    if failed_count > 0:
        print(f"❌ {failed_count} videos marked as permanently failed")
    
    save_video_status()

def enhanced_upload_video_to_api(video_file_path, organization_id=None, camera_guid=None, thumbnail_path=None):
    """Enhanced upload function with status tracking"""
    file_key = os.path.basename(video_file_path)
    
    try:
        # Update status to uploading
        update_video_status(video_file_path, 'uploading')
        
        # Call the original upload function
        success = upload_video_to_api(video_file_path, organization_id, camera_guid, thumbnail_path)
        
        if success:
            update_video_status(video_file_path, 'completed')
            print(f"✅ Successfully uploaded and marked as completed: {file_key}")
            return True
        else:
            # Increment retry count
            current_retry = video_status.get(file_key, {}).get('retry_count', 0)
            update_video_status(video_file_path, 'failed', 'Upload failed', current_retry + 1)
            print(f"❌ Upload failed, marked for retry: {file_key}")
            return False
            
    except Exception as e:
        current_retry = video_status.get(file_key, {}).get('retry_count', 0)
        update_video_status(video_file_path, 'failed', str(e), current_retry + 1)
        print(f"❌ Upload error: {file_key} - {e}")
        return False

class VideoFileHandler(FileSystemEventHandler):
    """File system event handler for video files"""
    
    def __init__(self, upload_callback):
        self.upload_callback = upload_callback
        self.pending_files = {}  # Track files being written
        
    def on_created(self, event):
        if not event.is_directory and event.src_path.endswith('.mp4'):
            print(f"📁 New video file detected: {os.path.basename(event.src_path)}")
            # Wait 30 seconds before even starting to monitor the file
            self.pending_files[event.src_path] = {
                'created_time': time.time(),
                'last_size': 0,
                'stable_count': 0,
                'monitoring_started': False
            }
    
    def on_modified(self, event):
        if not event.is_directory and event.src_path.endswith('.mp4'):
            if event.src_path in self.pending_files:
                self._check_file_stability(event.src_path)
    
    def _check_file_stability(self, file_path):
        """Check if file is stable (no longer being written) and has valid MP4 structure"""
        try:
            if not os.path.exists(file_path):
                return
                
            file_info = self.pending_files[file_path]
            current_time = time.time()
            
            # Wait 30 seconds before even starting to monitor the file
            if not file_info['monitoring_started']:
                if current_time - file_info['created_time'] < 30:
                    return  # Still waiting
                else:
                    file_info['monitoring_started'] = True
                    print(f"⏰ Starting to monitor file: {os.path.basename(file_path)}")
            
            current_size = os.path.getsize(file_path)
            
            # If file size hasn't changed, increment stable count
            if current_size == file_info['last_size'] and current_size > 0:
                file_info['stable_count'] += 1
            else:
                file_info['stable_count'] = 0
                file_info['last_size'] = current_size
            
            # If file has been stable for 15 consecutive checks (very conservative for MP4)
            if file_info['stable_count'] >= 15:
                # Additional MP4 structure validation
                if self._validate_mp4_structure(file_path):
                    print(f"✅ File stable and ready for upload: {os.path.basename(file_path)}")
                    del self.pending_files[file_path]
                    
                    # Check if file is not already processed
                    file_key = os.path.basename(file_path)
                    if file_key not in video_status or video_status[file_key]['status'] != 'completed':
                        update_video_status(file_path, 'pending')
                        self.upload_callback(file_path)
                    else:
                        print(f"📋 File already processed: {file_key}")
                else:
                    print(f"⚠️ File structure validation failed: {os.path.basename(file_path)}")
                    file_info['stable_count'] = 0  # Reset counter
                    
        except Exception as e:
            print(f"⚠️ Error checking file stability for {file_path}: {e}")
    
    def _validate_mp4_structure(self, file_path):
        """Validate MP4 file structure"""
        try:
            with open(file_path, 'rb') as f:
                # Read first 1024 bytes to check for MP4 header
                header = f.read(1024)
                if len(header) < 1024:
                    return False
                
                # Check for MP4 file signature (ftyp box)
                if b'ftyp' not in header and b'moov' not in header:
                    return False
                
                # Try to open with OpenCV to verify it's readable
                cap = cv2.VideoCapture(file_path)
                if not cap.isOpened():
                    return False
                
                # Try to read first frame
                ret, frame = cap.read()
                cap.release()
                
                return ret and frame is not None
                
        except Exception as e:
            print(f"⚠️ Error validating MP4 structure: {e}")
            return False

class VideoFileMonitor:
    """Monitors video directory for new files"""
    
    def __init__(self, watch_directory, upload_callback):
        self.watch_directory = watch_directory
        self.upload_callback = upload_callback
        self.observer = None
        self.event_handler = VideoFileHandler(upload_callback)
        self.monitor_thread = None
        self.running = False
    
    def start(self):
        """Start the file monitor"""
        if self.running:
            return
            
        self.running = True
        self.monitor_thread = threading.Thread(target=self._monitor_loop, daemon=True)
        self.monitor_thread.start()
        print(f"👀 Started video file monitor for: {self.watch_directory}")
    
    def stop(self):
        """Stop the file monitor"""
        self.running = False
        if self.observer:
            self.observer.stop()
            self.observer.join()
        print("🛑 Video file monitor stopped")
    
    def _monitor_loop(self):
        """Main monitoring loop"""
        try:
            self.observer = Observer()
            self.observer.schedule(self.event_handler, self.watch_directory, recursive=True)
            self.observer.start()
            
            while self.running:
                time.sleep(1)
                
        except Exception as e:
            print(f"❌ Error in file monitor: {e}")
        finally:
            if self.observer:
                self.observer.stop()
                self.observer.join()

# Global variables for dedicated upload thread
upload_thread = None
upload_thread_running = False
upload_queue = []  # Queue of videos waiting to be uploaded
uploaded_files = set()  # Set of files that have been uploaded

# Video state tracking for completion-based uploads
video_states = {}  # Track video states per camera GUID
last_video_times = {}  # Track last video modification times per camera

# Global file monitor
file_monitor = None

# Global cleanup thread
cleanup_thread = None
cleanup_thread_running = False

def cleanup_worker():
    """Dedicated cleanup thread that handles file management and uploads"""
    global cleanup_thread_running
    
    print("🧹 Cleanup worker thread started")
    cleanup_thread_running = True
    
    while cleanup_thread_running:
        try:
            # Process all camera directories
            if base_video_path and os.path.exists(base_video_path):
                for date_dir in os.listdir(base_video_path):
                    date_path = os.path.join(base_video_path, date_dir)
                    if not os.path.isdir(date_path):
                        continue
                    
                    # Look for camera GUID directories
                    for guid_dir in os.listdir(date_path):
                        guid_path = os.path.join(date_path, guid_dir)
                        if not os.path.isdir(guid_path):
                            continue
                        
                        # Process each hour directory
                        for hour_dir in os.listdir(guid_path):
                            hour_path = os.path.join(guid_path, hour_dir)
                            if not os.path.isdir(hour_path):
                                continue
                            
                            # Process video files in this hour directory
                            process_video_files_in_directory(hour_path, guid_dir)
            
            # Sleep before next cleanup cycle
            time.sleep(60)  # Check every 60 seconds to avoid processing incomplete files
            
        except Exception as e:
            print(f"❌ Error in cleanup worker: {e}")
            time.sleep(10)  # Shorter sleep on error
    
    print("🧹 Cleanup worker thread stopped")

def process_video_files_in_directory(directory_path, camera_guid):
    """Process video files in a specific directory for cleanup only (uploads handled by dedicated worker)"""
    try:
        # Get all .mp4 files in the directory
        video_files = [f for f in os.listdir(directory_path) if f.endswith('.mp4')]
        
        for video_file in video_files:
            file_path = os.path.join(directory_path, video_file)
            file_key = os.path.basename(file_path)
            
            # Check if file is already processed
            if file_key in video_status and video_status[file_key]['status'] == 'completed':
                continue
            
            # Check if file is stable and has minimum size
            if not is_file_stable(file_path):
                continue
            
            # Check minimum file size (1KB = 1024 bytes)
            file_size = os.path.getsize(file_path)
            if file_size < 1024:
                print(f"⚠️ File too small ({file_size} bytes), skipping: {video_file}")
                continue
            
            # Check file age (must be at least 30 seconds old for segmented MP4 files)
            try:
                file_age = time.time() - os.path.getctime(file_path)
                if file_age < 30:  # 30 seconds minimum age for segmented files
                    print(f"⚠️ File too new ({file_age:.1f}s), skipping: {video_file}")
                    continue
            except Exception as e:
                print(f"⚠️ Error checking file age: {e}")
                continue
            
            # Only cleanup old files, uploads are handled by dedicated upload worker
            print(f"🧹 Cleanup worker found stable file: {video_file} (upload handled by dedicated worker)")
                
    except Exception as e:
        print(f"❌ Error processing directory {directory_path}: {e}")

def is_file_stable(file_path, stability_checks=3):
    """Check if file is stable (not being written) and has valid MP4 structure"""
    try:
        if not os.path.exists(file_path):
            return False
        
        # Check if file is locked (being written to)
        try:
            with open(file_path, 'r+b') as f:
                pass  # If we can open in read-write mode, file is not locked
        except (PermissionError, OSError):
            print(f"⚠️ File is locked (being written): {os.path.basename(file_path)}")
            return False
        
        # Additional check: ensure file is not currently being written by checking if it's growing
        initial_size = os.path.getsize(file_path)
        time.sleep(1)  # Wait 1 second
        current_size = os.path.getsize(file_path)
        if current_size > initial_size:
            print(f"⚠️ File is still growing ({initial_size} -> {current_size}): {os.path.basename(file_path)}")
            return False
        
        # Check file size stability with shorter waits for faster processing
        size1 = os.path.getsize(file_path)
        time.sleep(2)  # Wait 2 seconds for MP4 files to be complete
        size2 = os.path.getsize(file_path)
        time.sleep(2)  # Wait another 2 seconds
        size3 = os.path.getsize(file_path)
        
        # File is stable if size hasn't changed for 3 checks
        if size1 != size2 or size2 != size3 or size1 <= 0:
            return False
        
        # Additional check: verify MP4 structure
        try:
            with open(file_path, 'rb') as f:
                # Read first 1024 bytes to check for MP4 header
                header = f.read(1024)
                if len(header) < 1024:
                    return False
                
                # Check for MP4 file signature (ftyp box)
                if b'ftyp' not in header and b'moov' not in header:
                    return False
                
                # Use ffprobe to verify file is playable (more reliable than OpenCV for segmented MP4)
                try:
                    import subprocess
                    result = subprocess.run(['ffprobe', '-v', 'quiet', '-print_format', 'json', '-show_format', file_path], 
                                          capture_output=True, text=True, timeout=5)
                    if result.returncode != 0:
                        print(f"⚠️ File not playable (ffprobe failed): {os.path.basename(file_path)}")
                        return False
                    
                    # Parse JSON to check duration and format
                    import json
                    probe_data = json.loads(result.stdout)
                    if 'format' not in probe_data or 'duration' not in probe_data['format']:
                        print(f"⚠️ Invalid format data: {os.path.basename(file_path)}")
                        return False
                        
                    duration = float(probe_data['format']['duration'])
                    if duration < 1.0:  # File too short, likely incomplete
                        print(f"⚠️ File too short ({duration}s): {os.path.basename(file_path)}")
                        return False
                        
                except Exception as e:
                    print(f"⚠️ Error testing file with ffprobe: {e}")
                    return False
                
                return True
                
        except Exception as e:
            print(f"⚠️ Error verifying MP4 structure for {os.path.basename(file_path)}: {e}")
            return False
        
    except Exception as e:
        print(f"⚠️ Error checking file stability: {e}")
        return False

def generate_video_thumbnail(file_path):
    """Generate thumbnail for video file"""
    try:
        thumbnail_filename = f"thumb_{os.path.splitext(os.path.basename(file_path))[0]}.jpg"
        thumbnail_dir = os.path.dirname(file_path)
        thumbnail_path = os.path.join(thumbnail_dir, thumbnail_filename)
        
        # Generate thumbnail using OpenCV
        cap = cv2.VideoCapture(file_path)
        if cap.isOpened():
            ret, frame = cap.read()
            if ret and frame is not None:
                # Resize frame for thumbnail
                height, width = frame.shape[:2]
                max_size = 320
                if width > height:
                    new_width = max_size
                    new_height = int(height * max_size / width)
                else:
                    new_height = max_size
                    new_width = int(width * max_size / height)
                
                thumbnail = cv2.resize(frame, (new_width, new_height))
                cv2.imwrite(thumbnail_path, thumbnail)
                cap.release()
                return thumbnail_path
            cap.release()
    except Exception as e:
        print(f"⚠️ Error generating thumbnail for {os.path.basename(file_path)}: {e}")
    
    return None

def start_cleanup_thread():
    """Start the cleanup thread"""
    global cleanup_thread, cleanup_thread_running
    
    if cleanup_thread_running:
        return
    
    cleanup_thread = threading.Thread(target=cleanup_worker, daemon=True, name="CleanupWorker")
    cleanup_thread.start()
    print("🧹 Cleanup thread started")

def stop_cleanup_thread():
    """Stop the cleanup thread"""
    global cleanup_thread_running
    
    cleanup_thread_running = False
    if cleanup_thread and cleanup_thread.is_alive():
        cleanup_thread.join(timeout=5)
    print("🧹 Cleanup thread stopped")

# video_upload_callback function removed - now using completion-based upload logic

# Known camera vendors to help classify devices discovered on the LAN
KNOWN_CAMERA_VENDORS = [
    'axis', 'hikvision', 'dahua', 'uniview', 'cp plus',
    'bosch', 'hanwha', 'sony', 'avtech', 'arecont', 'vivotek', 'mobotix',
]

def get_mac_vendor(mac_address):
    """Lookup vendor name for a MAC address using macvendors API.
    Returns 'Unknown' on error.
    """
    try:
        url = f"https://api.macvendors.com/{mac_address}"
        response = requests.get(url, timeout=5)
        if response.status_code == 200 and response.text:
            return response.text.strip()
    except Exception:
        pass
    return "Unknown"

def validate_video_file(video_path):
    """Comprehensive video file validation to prevent corrupted uploads"""
    try:
        # Check if file exists
        if not os.path.exists(video_path):
            print(f"❌ Video file not found: {video_path}")
            return False
        
        # Check file size
        file_size = os.path.getsize(video_path)
        if file_size < 1024:  # Less than 1KB
            print(f"❌ Video file too small ({file_size} bytes), likely corrupted")
            return False
        
        # Skip stability check - process files immediately
        
        # For segment files, we can't always expect complete MP4 structure
        # Try to open with OpenCV first as a better validation method
        try:
            cap = cv2.VideoCapture(video_path)
            if not cap.isOpened():
                print(f"❌ Cannot open video file with OpenCV: {video_path}")
                return False
            
            # Try to read first frame
            ret, frame = cap.read()
            if not ret or frame is None:
                print(f"❌ Cannot read frames from video file: {video_path}")
                cap.release()
                return False
            
            # Check video properties
            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            
            # For segment files, even 1 frame is acceptable
            if fps <= 0 or frame_count <= 0 or width <= 0 or height <= 0:
                print(f"❌ Invalid video properties: fps={fps}, frames={frame_count}, size={width}x{height}")
                cap.release()
                return False
            
            cap.release()
            
            print(f"✅ Video validation passed: {os.path.basename(video_path)} ({file_size} bytes, {frame_count} frames, {fps:.1f} fps)")
            return True
            
        except Exception as e:
            print(f"❌ Error validating video with OpenCV: {e}")
            # Fallback: check if file has basic video data
            try:
                with open(video_path, 'rb') as f:
                    header = f.read(8192)
                    if len(header) < 1024:
                        return False
                    # Look for any video data indicators
                    if b'mdat' in header or b'avc1' in header or b'h264' in header:
                        print(f"⚠️ Basic video validation passed (segment file): {os.path.basename(video_path)}")
                        return True
            except:
                pass
            return False
        
    except Exception as e:
        print(f"❌ Error in video validation: {e}")
        return False

def verify_upload_integrity(video_path, expected_size):
    """Verify that the video file is still intact after upload"""
    try:
        # Check if file still exists and has correct size
        if not os.path.exists(video_path):
            print(f"❌ Video file missing after upload: {video_path}")
            return False
        
        current_size = os.path.getsize(video_path)
        if current_size != expected_size:
            print(f"❌ File size changed after upload: expected {expected_size}, got {current_size}")
            return False
        
        # Quick validation that file can still be opened
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            print(f"❌ Cannot open video file after upload: {video_path}")
            return False
        
        # Try to read first frame
        ret, frame = cap.read()
        cap.release()
        
        if not ret or frame is None:
            print(f"❌ Cannot read frames after upload: {video_path}")
            return False
        
        print(f"✅ Upload integrity verified: {os.path.basename(video_path)}")
        return True
        
    except Exception as e:
        print(f"❌ Error verifying upload integrity: {e}")
        return False

def generate_video_thumbnail(video_path, thumbnail_path, timestamp_seconds=5):
    """Generate a thumbnail from video at specified timestamp"""
    try:
        # Use OpenCV to capture frame at specified timestamp
        cap = cv2.VideoCapture(video_path)
        
        if not cap.isOpened():
            print(f"❌ Could not open video for thumbnail: {video_path}")
            return False
        
        # Get video properties
        fps = cap.get(cv2.CAP_PROP_FPS)
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        
        # Calculate frame number for the timestamp
        frame_number = min(int(timestamp_seconds * fps), total_frames - 1)
        
        # Set frame position
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame_number)
        
        # Read frame
        ret, frame = cap.read()
        cap.release()
        
        if not ret or frame is None:
            print(f"❌ Could not read frame for thumbnail: {video_path}")
            return False
        
        # Resize frame to thumbnail size (320x240)
        height, width = frame.shape[:2]
        aspect_ratio = width / height
        new_width = 320
        new_height = int(new_width / aspect_ratio)
        
        resized_frame = cv2.resize(frame, (new_width, new_height))
        
        # Convert BGR to RGB for PIL
        rgb_frame = cv2.cvtColor(resized_frame, cv2.COLOR_BGR2RGB)
        
        # Convert to PIL Image and save
        pil_image = Image.fromarray(rgb_frame)
        pil_image.save(thumbnail_path, 'JPEG', quality=85)
        
        print(f"✅ Thumbnail generated: {os.path.basename(thumbnail_path)}")
        return True
        
    except Exception as e:
        print(f"❌ Error generating thumbnail for {video_path}: {e}")
        return False

def dedicated_upload_worker():
    """Dedicated thread that waits for video completion before uploading"""
    global upload_thread_running, uploaded_files, video_states, last_video_times
    
    print("🚀 Starting dedicated upload worker thread...")
    
    while upload_thread_running:
        try:
            # Scan for new video files in the videos directory
            if os.path.exists(base_video_path):
                current_time = time.time()
                
                # Get all video files sorted by modification time
                video_files = []
                for root, dirs, files in os.walk(base_video_path):
                    for file in files:
                        if file.endswith('.mp4'):
                            file_path = os.path.join(root, file)
                            if file_path not in uploaded_files:
                                try:
                                    mod_time = os.path.getmtime(file_path)
                                    video_files.append((file_path, mod_time))
                                except OSError:
                                    continue
                
                # Sort by modification time (oldest first)
                video_files.sort(key=lambda x: x[1])
                
                for file_path, mod_time in video_files:
                    try:
                        # Extract camera GUID from path
                        path_parts = file_path.replace('\\', '/').split('/')
                        camera_guid = None
                        
                        # Look for GUID in path (format: videos/date/guid/hour/file.mp4)
                        for i, part in enumerate(path_parts):
                            if part == 'videos' and i + 3 < len(path_parts):
                                # Check if next part looks like a date (YYYY-MM-DD)
                                if len(path_parts[i + 1]) == 10 and '-' in path_parts[i + 1]:
                                    camera_guid = path_parts[i + 2]
                                    break
                        
                        if not camera_guid:
                            print(f"⚠️ Could not extract camera GUID from path: {file_path}")
                            uploaded_files.add(file_path)
                            continue
                        
                        # Initialize camera state if not exists
                        if camera_guid not in video_states:
                            video_states[camera_guid] = {
                                'current_file': None,
                                'last_upload_time': 0,
                                'pending_uploads': []
                            }
                        
                        camera_state = video_states[camera_guid]
                        
                        # Check if this is a new video file (different from current)
                        if camera_state['current_file'] != file_path:
                            # If there was a previous file, it's now complete - add to pending uploads
                            if camera_state['current_file'] and camera_state['current_file'] not in uploaded_files:
                                print(f"📹 Video completed for camera {camera_guid}: {os.path.basename(camera_state['current_file'])}")
                                camera_state['pending_uploads'].append(camera_state['current_file'])
                            
                            # Update current file
                            camera_state['current_file'] = file_path
                            last_video_times[camera_guid] = mod_time
                            print(f"📹 New video started for camera {camera_guid}: {os.path.basename(file_path)}")
                        
                        # Check if current file is stable (not being written to)
                        if camera_state['current_file'] == file_path:
                            if is_file_stable(file_path, stability_checks=3):
                                # File is stable, but don't upload yet - wait for next video
                                print(f"✅ Video file stable for camera {camera_guid}: {os.path.basename(file_path)} (waiting for next video)")
                            else:
                                print(f"⏳ Video file still being written for camera {camera_guid}: {os.path.basename(file_path)}")
                        
                        # Process pending uploads (completed videos)
                        if camera_state['pending_uploads']:
                            print(f"📤 Processing {len(camera_state['pending_uploads'])} pending uploads for camera {camera_guid}")
                            
                            for pending_file in camera_state['pending_uploads'][:]:  # Copy list to avoid modification during iteration
                                try:
                                    # Double-check file is still stable before upload
                                    if is_file_stable(pending_file, stability_checks=1):
                                        print(f"🚀 Uploading completed video: {os.path.basename(pending_file)}")
                                        
                                        # Generate thumbnail
                                        thumbnail_path = None
                                        try:
                                            thumbnail_filename = f"thumb_{os.path.splitext(os.path.basename(pending_file))[0]}.jpg"
                                            thumbnail_path = os.path.join(os.path.dirname(pending_file), thumbnail_filename)
                                            
                                            print(f"📷 Generating thumbnail for {os.path.basename(pending_file)}...")
                                            if generate_video_thumbnail(pending_file, thumbnail_path, timestamp_seconds=5):
                                                print(f"✅ Thumbnail generated: {thumbnail_filename}")
                                            else:
                                                print(f"⚠️ Thumbnail generation failed for {os.path.basename(pending_file)}")
                                                thumbnail_path = None
                                        except Exception as e:
                                            print(f"⚠️ Error generating thumbnail for {os.path.basename(pending_file)}: {e}")
                                            thumbnail_path = None
                                        
                                        # Upload video
                                        upload_success = upload_video_to_api(
                                            pending_file,
                                            organization_id=api_organization_id,
                                            camera_guid=camera_guid,
                                            thumbnail_path=thumbnail_path
                                        )
                                        
                                        if upload_success:
                                            print(f"✅ Successfully uploaded {os.path.basename(pending_file)}")
                                            uploaded_files.add(pending_file)
                                            camera_state['last_upload_time'] = current_time
                                            
                                            # Clean up thumbnail
                                            if thumbnail_path and os.path.exists(thumbnail_path):
                                                try:
                                                    os.remove(thumbnail_path)
                                                    print(f"🗑️ Cleaned up thumbnail: {os.path.basename(thumbnail_path)}")
                                                except Exception as e:
                                                    print(f"⚠️ Could not clean up thumbnail: {e}")
                                        else:
                                            print(f"❌ Failed to upload {os.path.basename(pending_file)}")
                                        
                                        # Remove from pending list
                                        camera_state['pending_uploads'].remove(pending_file)
                                        
                                    else:
                                        print(f"⏳ Pending file still being written: {os.path.basename(pending_file)}")
                                        
                                except Exception as e:
                                    print(f"⚠️ Error processing pending file {os.path.basename(pending_file)}: {e}")
                                    # Remove from pending list to avoid infinite retry
                                    camera_state['pending_uploads'].remove(pending_file)
                    
                    except Exception as e:
                        print(f"⚠️ Error processing file {os.path.basename(file_path)}: {e}")
                        continue
            
            # Sleep before next scan
            time.sleep(10)  # Scan every 10 seconds
            
        except Exception as e:
            print(f"❌ Error in dedicated upload worker: {e}")
            time.sleep(15)  # Wait longer on error
    
    print("🛑 Dedicated upload worker thread stopped")

def start_dedicated_upload_thread():
    """Start the dedicated upload thread"""
    global upload_thread, upload_thread_running
    
    if upload_thread_running:
        print("⚠️ Dedicated upload thread is already running")
        return
    
    upload_thread_running = True
    upload_thread = threading.Thread(target=dedicated_upload_worker, name="DedicatedUploadWorker")
    upload_thread.daemon = True
    upload_thread.start()
    print("✅ Dedicated upload thread started")

def stop_dedicated_upload_thread():
    """Stop the dedicated upload thread"""
    global upload_thread_running, upload_thread
    
    if not upload_thread_running:
        print("⚠️ Dedicated upload thread is not running")
        return
    
    upload_thread_running = False
    if upload_thread and upload_thread.is_alive():
        upload_thread.join(timeout=5)
    print("✅ Dedicated upload thread stopped")

def test_rtsp_stream(camera_rtsp_url, camera_name):
    """Test if RTSP stream is accessible using ffmpeg instead of OpenCV"""
    try:
        print(f"Testing RTSP stream for {camera_name}: {camera_rtsp_url}")
        
        # Use ffmpeg to test the RTSP stream instead of OpenCV
        # This is more reliable for RTSP streams
        ffmpeg_cmd = [
            "ffmpeg",
            "-rtsp_transport", "tcp",
            "-i", camera_rtsp_url,
            "-t", "5",  # Test for 5 seconds
            "-f", "null",  # No output file needed
            "-"
        ]
        
        print(f"Running RTSP test command: {' '.join(ffmpeg_cmd)}")
        result = subprocess.run(
            ffmpeg_cmd,
            capture_output=True,
            text=True,
            timeout=10  # 10 second timeout
        )
        
        if result.returncode == 0:
            print(f"✅ RTSP stream test successful for {camera_name}")
            return True
        else:
            print(f"❌ RTSP stream test failed for {camera_name}")
            print(f"FFmpeg error: {result.stderr}")
            return False
            
    except subprocess.TimeoutExpired:
        print(f"❌ RTSP stream test timed out for {camera_name}")
        return False
    except Exception as e:
        print(f"❌ Error testing RTSP stream for {camera_name}: {e}")
        # If ffmpeg test fails, try a simple connectivity test
        try:
            import socket
            from urllib.parse import urlparse
            parsed = urlparse(camera_rtsp_url)
            host = parsed.hostname
            port = parsed.port or 554
            
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(5)
            result = sock.connect_ex((host, port))
            sock.close()
            
            if result == 0:
                print(f"✅ Camera {camera_name} is reachable on port {port}, proceeding with recording")
                return True
            else:
                print(f"❌ Camera {camera_name} is not reachable on port {port}")
                return False
        except Exception as e2:
            print(f"❌ Connectivity test also failed: {e2}")
            return False

def cleanup_corrupted_files(output_dir, camera_name):
    """Clean up corrupted or incomplete video files"""
    try:
        if not os.path.exists(output_dir):
            return
        
        corrupted_files = []
        for file in os.listdir(output_dir):
            if file.endswith('.mp4'):
                file_path = os.path.join(output_dir, file)
                try:
                    file_size = os.path.getsize(file_path)
                    
                    # Check if file is too small
                    if file_size < 1024:
                        corrupted_files.append((file_path, f"Too small ({file_size} bytes)"))
                        continue
                    
                    # Check if file can be read and has valid MP4 structure
                    try:
                        with open(file_path, 'rb') as f:
                            header = f.read(1024)
                            if len(header) < 1024 or b'mdat' not in header:
                                corrupted_files.append((file_path, "Invalid MP4 structure"))
                                continue
                    except Exception as e:
                        corrupted_files.append((file_path, f"Cannot read file: {e}"))
                        continue
                        
                except Exception as e:
                    corrupted_files.append((file_path, f"Error checking file: {e}"))
        
        # Remove corrupted files
        for file_path, reason in corrupted_files:
            try:
                # Try to remove the file, but handle "file in use" errors gracefully
                os.remove(file_path)
                print(f"🗑️ Cleaned up corrupted file [{camera_name}]: {os.path.basename(file_path)} - {reason}")
            except PermissionError as e:
                if "being used by another process" in str(e):
                    print(f"⏳ File {os.path.basename(file_path)} is currently being used by FFmpeg, will retry later")
                else:
                    print(f"⚠️ Permission error removing {file_path}: {e}")
            except Exception as e:
                print(f"⚠️ Failed to remove corrupted file {file_path}: {e}")
        
        if corrupted_files:
            print(f"🧹 Cleaned up {len(corrupted_files)} corrupted files for camera {camera_name}")
        
    except Exception as e:
        print(f"Error in cleanup_corrupted_files for {camera_name}: {e}")

def upload_video_to_api(video_file_path, organization_id=None, camera_guid=None, thumbnail_path=None):
    """Upload video file to API with optional thumbnail"""
    try:
        print(f"🔍 Attempting to upload: {os.path.basename(video_file_path)}")
        
        # Skip health check since API server doesn't have a health endpoint
        # The actual upload attempt will determine if the server is working
        print(f"📡 Proceeding with upload to API server...")
        
        # Skip video validation - push directly to API
        print(f"📤 Skipping validation, pushing directly to API: {os.path.basename(video_file_path)}")
        
        # Upload to API if configured
        if video_config['upload_to_api']:
            try:
                # Get file size for integrity check
                file_size = os.path.getsize(video_file_path)
                
                # Prepare files for upload - read files into memory to avoid file handle issues
                try:
                    # Read video file into memory
                    with open(video_file_path, 'rb') as video_file:
                        video_data = video_file.read()
                    
                    # Read thumbnail file into memory if available
                    thumbnail_data = None
                    if thumbnail_path and os.path.exists(thumbnail_path):
                        with open(thumbnail_path, 'rb') as thumb_file:
                            thumbnail_data = thumb_file.read()
                        print(f"📷 Including thumbnail: {os.path.basename(thumbnail_path)}")
                except Exception as e:
                    print(f"❌ Error reading files for upload: {e}")
                    return False
                
                # Prepare form data - always use organization ID from config as source of truth
                data = {
                    'organization_id': str(api_organization_id),  # Always use config organization ID
                    'guid': str(camera_guid) if camera_guid else str(uuid.uuid4())
                }
                
                # Log upload attempt with more details
                print(f"📤 Starting upload attempt for {os.path.basename(video_file_path)}")
                print(f"📤 Target API: {video_config['api_url']}")
                print(f"📤 Organization ID: {data['organization_id']}")
                print(f"📤 Camera GUID: {data['guid']}")
                print(f"📤 File size: {file_size} bytes")
                
                # Upload with retry logic
                max_retries = video_config.get('upload_retry_attempts', 3)
                upload_timeout = video_config.get('upload_timeout', 30)
                
                for attempt in range(max_retries):
                    try:
                        # Create files dict with BytesIO objects
                        files = {
                            'video_file': (os.path.basename(video_file_path), BytesIO(video_data), 'video/mp4')
                        }
                        
                        if thumbnail_data:
                            files['thumbnail_file'] = (os.path.basename(thumbnail_path), BytesIO(thumbnail_data), 'image/jpeg')
                        
                        response = requests.post(
                            video_config['api_url'],
                            files=files,
                            data=data,
                            timeout=upload_timeout
                        )
                        
                        if response.status_code == 200:
                            result = response.json()
                            print(f"✅ Video uploaded successfully: {result.get('message', 'Upload completed')}")
                            print(f"✅ File path: {result.get('file_path', 'Unknown')}")
                            
                            # Verify upload integrity by checking server response
                            if result.get('file_size') and int(result.get('file_size', 0)) != file_size:
                                print(f"⚠️ File size mismatch! Local: {file_size}, Server: {result.get('file_size')}")
                                return False
                            
                            print(f"✅ Upload integrity verified (size: {file_size} bytes)")
                            
                            # Additional integrity check after upload
                            if not verify_upload_integrity(video_file_path, file_size):
                                print(f"⚠️ Post-upload integrity check failed for {os.path.basename(video_file_path)}")
                                return False
                            
                            return True
                        else:
                            print(f"❌ Upload failed (attempt {attempt + 1}/{max_retries}): {response.status_code} - {response.text}")
                            if attempt < max_retries - 1:
                                time.sleep(2 ** attempt)  # Exponential backoff
                            
                    except requests.exceptions.RequestException as e:
                        print(f"❌ Upload error (attempt {attempt + 1}/{max_retries}): {e}")
                        if attempt < max_retries - 1:
                            time.sleep(2 ** attempt)  # Exponential backoff
                
                print(f"❌ All upload attempts failed for {os.path.basename(video_file_path)}")
                return False
                    
            except Exception as e:
                print(f"❌ Error uploading to API: {e}")
                return False
        else:
            # API upload is disabled, just store locally
            print(f"✅ Video stored locally (API disabled): {os.path.basename(video_file_path)} ({file_size} bytes)")
            print(f"📁 Local path: {video_file_path}")
            return True
        
    except Exception as e:
        print(f"❌ Error in upload_video_to_api: {e}")
        return False

def start_hourly_merging(camera_guid):
    """Start the hourly merging task for a camera"""
    def merging_worker():
        while True:
            try:
                # Wait until the next hour starts
                now = datetime.now()
                next_hour = now.replace(minute=0, second=0, microsecond=0)
                if next_hour <= now:
                    next_hour = next_hour.replace(hour=next_hour.hour + 1)
                
                # Calculate seconds to wait
                wait_seconds = (next_hour - now).total_seconds()
                print(f"Next merging task for camera {camera_guid} in {wait_seconds:.0f} seconds")
                
                # Wait until next hour
                time.sleep(wait_seconds)
                
                # Get the hour that just completed (previous hour)
                # Handle hour rollover properly (e.g., if next_hour is 0, previous hour is 23)
                if next_hour.hour == 0:
                    completed_hour = "23"
                else:
                    completed_hour = (next_hour.replace(hour=next_hour.hour - 1)).strftime("%H")
                
                # For now, just log the merge task
                print(f"Hourly merge task completed for hour {completed_hour} in camera {camera_guid}")
                
            except Exception as e:
                print(f"Error in merging worker for camera {camera_guid}: {e}")
                time.sleep(60)  # Wait a minute before retrying
    
    # Start merging thread
    merging_thread = threading.Thread(target=merging_worker)
    merging_thread.daemon = True
    merging_tasks[camera_guid] = merging_thread
    merging_thread.start()
    print(f"Started hourly merging task for camera {camera_guid}")

def find_lan_devices():
    """Discover devices using ARP table. Returns list of (ip, mac, vendor)."""
    try:
        output = os.popen("arp -a").read()
    except Exception:
        output = ""

    devices = re.findall(r"(\d+\.\d+\.\d+\.\d+)\s+([0-9a-f\-:]{17})", output, re.I)

    results = []
    for ip_address, mac in devices:
        normalized_mac = mac.replace('-', ':').lower()
        vendor = get_mac_vendor(normalized_mac)
        results.append((ip_address, normalized_mac, vendor))
    return results

def camera_thread_function(camera_id, camera_name, camera_rtsp_url, camera_guid):
    """Function that runs in each camera thread - creates video recordings using ffmpeg"""
    print(f"Thread started for camera: {camera_name} ({camera_id})")
    
    # Initialize hour tracking variables
    last_hour = None
    current_date = None
    current_hour = None
    output_dir = None
    
    def update_output_directory():
        """Update output directory when hour changes"""
        nonlocal current_date, current_hour, output_dir, last_hour
        
        now = datetime.now()
        new_date = now.strftime("%Y-%m-%d")
        new_hour = now.strftime("%H")
        
        # Check if date or hour has changed
        if current_date != new_date or current_hour != new_hour:
            current_date = new_date
            current_hour = new_hour
            output_dir = os.path.join(base_video_path, current_date, camera_guid, current_hour)
            os.makedirs(output_dir, exist_ok=True)
            
            if last_hour is not None:
                print(f"Hour changed from {last_hour} to {current_hour} for camera {camera_name}")
                print(f"New output directory: {output_dir}")
            
            last_hour = current_hour
            return True  # Directory was updated
        return False  # No change
    
    # Initialize the first output directory
    update_output_directory()
    
    # Start hourly merging task
    start_hourly_merging(camera_guid)
    
    try:
        print(f"Starting recording for camera: {camera_name} ({camera_id})")
        
        # Keep trying to start FFmpeg until stop flag is set
        while not stop_flags.get(camera_id, False):
            try:
                # Check if hour has changed and update output directory
                hour_changed = update_output_directory()
                if hour_changed:
                    print(f"Hour changed for camera {camera_name}, will restart recording in new directory")
                
                # Test RTSP stream (with fallback option)
                if not test_rtsp_stream(camera_rtsp_url, camera_name):
                    print(f"⚠️ RTSP stream test failed for camera: {camera_name} ({camera_id})")
                    print(f"⚠️ Proceeding with recording anyway - FFmpeg may still be able to connect")
                    # Don't break here - let FFmpeg try to connect directly
                
                # Get camera settings from config
                camera_config = None
                for cam in cameras_data:
                    if cam['id'] == camera_id:
                        camera_config = cam
                        break
                
                # Default settings
                frame_rate = 15
                video_bitrate = "110k"
                audio_bitrate = "24k"
                resolution = "1280:720"
                preset = "veryfast"
                crf = "28"
                segment_time = 60
                enable_audio = True
                
                # Override with camera-specific settings if available
                if camera_config:
                    if 'camera_settings' in camera_config:
                        settings = camera_config['camera_settings']
                        frame_rate = settings.get('frame_rate', frame_rate)
                        video_bitrate = f"{settings.get('video_bitrate', 110)}k"
                        audio_bitrate = f"{settings.get('audio_bitrate', 24)}k"
                        preset = settings.get('preset', preset)
                        crf = str(settings.get('crf', crf))
                        segment_time = settings.get('segment_time', segment_time)
                        enable_audio = settings.get('enable_audio', enable_audio)
                        
                        # Handle resolution
                        res = settings.get('resolution', '1280x720')
                        if 'x' in res:
                            width, height = res.split('x')
                            resolution = f"{width}:{height}"
                
                # Video recording command
                video_ffmpeg_cmd = [
            "ffmpeg",
            "-rtsp_transport", "tcp",
                    "-i", camera_rtsp_url,
                    "-c:v", "libx264",
                    "-b:v", video_bitrate,
                    "-maxrate", video_bitrate,
                    "-bufsize", f"{int(video_bitrate.replace('k', '')) * 2}k",
                    "-preset", preset,
                    "-crf", crf,
                    "-r", str(frame_rate),
                    "-vf", f"scale={resolution}",
        "-avoid_negative_ts", "make_zero",  # Handle timestamp issues
                    "-fflags", "+genpts"
                ]
                
                # Add audio settings if enabled
                if enable_audio:
                    video_ffmpeg_cmd.extend([
                        "-c:a", "aac",
                        "-b:a", audio_bitrate,
                    ])
                
                # Create complete MP4 files with proper segmentation
                video_ffmpeg_cmd.extend([
            "-f", "segment",
                    "-segment_format", "mp4",
                    "-movflags", "+faststart",  # Create complete MP4 files
                    "-segment_time", str(segment_time),
                    "-segment_time_delta", "0.1",
            "-reset_timestamps", "1",
                    "-segment_atclocktime", "1",  # Align to clock time
                    "-strftime", "1",  # Enable strftime in segment names
                    "-force_key_frames", f"expr:gte(t,n_forced*{segment_time})",  # Force keyframes at segment boundaries
                    "-g", str(int(frame_rate) * segment_time),  # Set GOP size to match segment time
            os.path.join(output_dir, "%H%M.mp4")
                ])
                
                print(f"Running video ffmpeg command: {' '.join(video_ffmpeg_cmd)}")
                
                # Start video ffmpeg process
                video_process = subprocess.Popen(
                    video_ffmpeg_cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
                    universal_newlines=True,
                    bufsize=1
                )
                
                print(f"Video FFmpeg process started with PID: {video_process.pid}")
                
                # Monitor FFmpeg output in real-time
                def monitor_video_ffmpeg():
                    while video_process.poll() is None:
                        line = video_process.stderr.readline()
                        if line:
                            print(f"Video FFmpeg [{camera_name}]: {line.strip()}")
                
                # Start monitoring thread
                video_monitor_thread = threading.Thread(target=monitor_video_ffmpeg)
                video_monitor_thread.daemon = True
                video_monitor_thread.start()
                
                # Simple file monitoring for logging only (no upload processing)
                def monitor_files():
                    last_files = set()
                    while video_process.poll() is None and not stop_flags.get(camera_id, False):
                        try:
                            # Update output directory in case hour changed
                            update_output_directory()
                            
                            current_files = set()
                            if os.path.exists(output_dir):
                                for file in os.listdir(output_dir):
                                    if file.endswith('.mp4'):
                                        current_files.add(file)
                            
                            # Check for new files and log them
                            new_files = current_files - last_files
                            if new_files:
                                for new_file in new_files:
                                    print(f"📹 New video file created [{camera_name}]: {new_file}")
                            
                            last_files = current_files
                            time.sleep(5)  # Check every 5 seconds
                            
                        except Exception as e:
                            print(f"Error monitoring files for {camera_name}: {e}")
                            time.sleep(5)
                
                # Start simple file monitoring thread
                file_monitor_thread = threading.Thread(target=monitor_files)
                file_monitor_thread.daemon = True
                file_monitor_thread.start()
            
                # Wait for process to complete or stop flag to be set
                while video_process.poll() is None and not stop_flags.get(camera_id, False):
                    # Check if hour has changed during recording
                    hour_changed = update_output_directory()
                    if hour_changed:
                        print(f"Hour changed during recording for camera {camera_name}, restarting FFmpeg...")
                        # Terminate current process
                        if video_process.poll() is None:
                            video_process.terminate()
                            try:
                                video_process.wait(timeout=5)
                            except subprocess.TimeoutExpired:
                                video_process.kill()
                        break  # Break out of this loop to restart FFmpeg with new directory
                    time.sleep(1)
                
                # If stop flag is set, terminate the process
                if stop_flags.get(camera_id, False):
                    print(f"Stopping recording for camera: {camera_name} ({camera_id})")
                    # Terminate video process
                    if video_process.poll() is None:
                        video_process.terminate()
                        try:
                            video_process.wait(timeout=5)
                        except subprocess.TimeoutExpired:
                            video_process.kill()
                    break
                
                # If process ended naturally, log it and restart
                if video_process.poll() is not None and not stop_flags.get(camera_id, False):
                    print(f"Recording process ended unexpectedly for camera: {camera_name} ({camera_id})")
                    # Check video process
                    stdout, stderr = video_process.communicate()
                    if stderr:
                        print(f"Video FFmpeg error output: {stderr}")
                    if stdout:
                        print(f"Video FFmpeg output: {stdout}")
                    print(f"Restarting FFmpeg for camera: {camera_name} ({camera_id}) in 5 seconds...")
                    time.sleep(5)
            
            except Exception as e:
                print(f"Error in FFmpeg process for camera {camera_name} ({camera_id}): {e}")
                print(f"Restarting FFmpeg for camera: {camera_name} ({camera_id}) in 5 seconds...")
                time.sleep(5)
            
    except Exception as e:
        print(f"Error in recording for camera {camera_name} ({camera_id}): {e}")
    
    print(f"Thread stopped for camera: {camera_name} ({camera_id})")

def start_camera_thread(camera):
    """Start a thread for a specific camera"""
    camera_id = camera['id']
    camera_name = camera['name']
    camera_url = camera.get('url', camera.get('URL', ''))  # Get URL from camera config (external API uses lowercase)
    camera_guid = camera.get('guid', camera.get('GUID', camera_id))  # Get GUID from camera config (external API uses lowercase)
    
    if not camera_url:
        print(f"No URL found for camera: {camera_name} ({camera_id})")
        return
    
    # If thread already exists, stop it first
    if camera_id in active_threads:
        stop_camera_thread(camera_id)
    
    # Set stop flag to False
    stop_flags[camera_id] = False
    
    # Create and start new thread
    thread = threading.Thread(target=camera_thread_function, args=(camera_id, camera_name, camera_url, camera_guid))
    thread.daemon = True  # Make thread daemon so it stops when main program exits
    active_threads[camera_id] = thread
    thread.start()
    
    print(f"Started recording thread for camera: {camera_name}")

def stop_camera_thread(camera_id):
    """Stop a thread for a specific camera"""
    if camera_id in active_threads:
        # Set stop flag to True
        stop_flags[camera_id] = True
        
        # Wait for thread to finish (with timeout)
        thread = active_threads[camera_id]
        thread.join(timeout=5)  # Wait up to 5 seconds
        
        # Remove from active threads
        del active_threads[camera_id]
        if camera_id in stop_flags:
            del stop_flags[camera_id]
        
        print(f"Stopped recording thread for camera ID: {camera_id}")

def validate_camera_credentials(camera):
    """Validate that camera credentials match the RTSP URL"""
    try:
        username = camera.get('username', 'root')
        password = camera.get('password', '')
        ip_address = camera.get('ip_address', '')
        port = camera.get('port', '554')
        path = camera.get('path', '/axis-media/media.amp')
        
        # Construct expected RTSP URL
        if username and password and ip_address:
            encoded_password = urllib.parse.quote(password, safe='')
            # Only include port if it's not the default RTSP port (554) or if it's explicitly set
            if port and port != '554' and port != '':
                expected_url = f"rtsp://{username}:{encoded_password}@{ip_address}:{port}{path}"
            else:
                expected_url = f"rtsp://{username}:{encoded_password}@{ip_address}{path}"
            
            # Check if the stored URL matches the constructed one
            stored_url = camera.get('url') or camera.get('URL', '')
            if stored_url and stored_url != expected_url:
                print(f"Warning: RTSP URL mismatch for camera {camera.get('name', 'Unknown')}")
                print(f"Expected: {expected_url}")
                print(f"Stored: {stored_url}")
                # Update the stored URL to match credentials
                camera['url'] = expected_url
                camera['URL'] = expected_url
                return True
        return True
    except Exception as e:
        print(f"Error validating credentials for camera {camera.get('name', 'Unknown')}: {e}")
        return False

def load_cameras():
    """Load cameras from config.json"""
    global cameras_data, base_video_path, groups_data, external_api_config, server_config
    try:
        with open('config.json', 'r') as f:
            config = json.load(f)
        
        cameras_data = config.get('cameras', [])
        base_video_path = config.get('base_video_path', './videos')
        groups_data = config.get('groups', [])
        external_api_config = config.get('external_api', external_api_config)
        server_config = config.get('server', server_config)
        
        # Create base video directory if it doesn't exist
        os.makedirs(base_video_path, exist_ok=True)
        
        # Validate and sync camera credentials
        for camera in cameras_data:
            validate_camera_credentials(camera)
        
        print(f"Loaded {len(cameras_data)} cameras from config.json")
        print(f"Base video path: {base_video_path}")
        print(f"External API config: {get_api_base_url()}")
        print(f"Server config: {server_config['host']}:{server_config['port']} (debug: {server_config['debug']})")
        print(f"Loaded {len(groups_data)} groups from config.json")
        
        # Start recording for cameras that were already recording
        for camera in cameras_data:
            if camera.get('is_recording', False):
                start_camera_thread(camera)
                
    except FileNotFoundError:
        print("config.json not found. Creating default config...")
        cameras_data = []
        base_video_path = './videos'
        groups_data = []
        os.makedirs(base_video_path, exist_ok=True)
        
        # Create default config
        default_config = {
            'base_video_path': base_video_path,
            'cameras': [],
            'groups': []
        }
        
        with open('config.json', 'w') as f:
            json.dump(default_config, f, indent=2)
            
    except Exception as e:
        print(f"Error loading cameras: {e}")
        cameras_data = []
        base_video_path = './videos'

def save_cameras():
    """Save cameras and groups back to config.json"""
    try:
        with open('config.json', 'r') as f:
            config = json.load(f)
        
        config['cameras'] = cameras_data
        config['groups'] = groups_data
        
        with open('config.json', 'w') as f:
            json.dump(config, f, indent=2)
        return True
    except Exception as e:
        print(f"Error saving cameras: {e}")
        return False

@app.route('/')
def index():
    """Serve the main HTML page"""
    return send_from_directory('.', 'index.html')

@app.route('/api/external-api-config')
def get_external_api_config():
    """API endpoint to get external API configuration"""
    return jsonify(external_api_config)

@app.route('/api/cameras')
def get_cameras():
    """API endpoint to get all cameras"""
    return jsonify(cameras_data)

@app.route('/api/cameras/fetch', methods=['POST'])
def fetch_cameras():
    """API endpoint to fetch cameras from external API and update config.json"""
    try:
        data = request.get_json() or {}
        organization_id = data.get('organization_id')
        
        result = fetch_cameras_from_api(organization_id)
        
        if result is not None:
            return jsonify({
                'message': 'Cameras fetched successfully',
                'count': len(result),
                'cameras': result
            })
        else:
            return jsonify({'error': 'Failed to fetch cameras from API'}), 500
            
    except Exception as e:
        print(f"Error in fetch_cameras endpoint: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cameras', methods=['POST'])
def create_camera():
    """Create a new camera and persist in config.json
    Expected JSON: {id?, name, location, ip_address, URL/url, GUID?, is_recording?, 
                   frame_rate?, resolution?, video_bitrate?, enable_audio?, audio_bitrate?, 
                   preset?, crf?, segment_time?, detection_features?, alert_settings?}
    """
    global cameras_data
    data = request.get_json(force=True) or {}

    # Basic validation
    name = (data.get('name') or '').strip()
    ip_address = (data.get('ip_address') or '').strip()
    location = (data.get('location') or '').strip()
    url_value = data.get('URL') or data.get('url') or ''
    url_value = url_value.strip()
    if not name or not ip_address:
        return jsonify({'error': 'name and ip_address are required'}), 400

    # Prevent duplicate id/ip
    if any(c.get('ip_address') == ip_address for c in cameras_data):
        return jsonify({'error': 'Camera with this IP already exists'}), 409

    camera_id = data.get('id') or f"cam_{uuid.uuid4().hex[:8]}"
    if any(c.get('id') == camera_id for c in cameras_data):
        return jsonify({'error': 'Camera with this ID already exists'}), 409

    guid = data.get('GUID') or str(uuid.uuid4())
    is_recording = bool(data.get('is_recording', False))
    status = data.get('status') or 'Active'

    # Default camera settings
    default_camera_settings = {
        'frame_rate': 30,
        'resolution': '1920x1080',
        'video_bitrate': 4000,
        'enable_audio': True,
        'audio_bitrate': 128,
        'preset': 'medium',
        'crf': 23,
        'segment_time': 60
    }

    # Default detection features (all false)
    default_detection_features = {
        'people_detection': False,
        'vehicle_detection': False,
        'people_counting': False,
        'vehicle_counting': False,
        'face_detection': False,
        'face_recognition': False,
        'mask_detection': False,
        'face_mask_compliance': False,
        'intrusion_detection': False,
        'line_crossing_detection': False,
        'loitering_detection': False,
        'aggressive_behavior_detection': False,
        'suspicious_behavior_detection': False,
        'fall_detection': False,
        'social_distancing_detection': False,
        'object_left_behind': False,
        'object_removed': False,
        'bag_detection': False,
        'gun_detection': False,
        'helmet_detection': False,
        'ppe_detection': False,
        'smoking_detection': False,
        'phone_usage_detection': False,
        'license_plate_recognition': False,
        'parking_occupancy_detection': False,
        'wrong_way_detection': False,
        'speed_detection': False,
        'illegal_parking_detection': False,
        'crowd_density_estimation': False,
        'dwell_time_analysis': False,
        'heatmap_generation': False,
        'queue_monitoring': False,
        'footfall_analysis': False,
        'wait_time_analysis': False,
        'fire_detection': False,
        'smoke_detection': False,
        'flood_water_level_detection': False,
        'night_time_movement_detection': False,
        'scream_gunshot_detection': False,
        'glass_break_detection': False,
        're_identification': False,
        'multi_camera_tracking': False,
        'pose_estimation': False,
        'action_recognition': False,
        'gesture_recognition': False,
        'scene_anomaly_detection': False,
        'zone_based_alerts': False,
        'virtual_tripwire': False,
        'retail_zone_interaction': False
    }

    # Default alert settings
    default_alert_settings = {
        'real_time_alerts': False,
        'sms_alerts': False,
        'email_alerts': False,
        'in_app_alerts': True,
        'phone_numbers': [],
        'email_addresses': [],
        'alert_sensitivity': 'Medium',
        'notification_cooldown': 5
    }

    # Merge with provided data
    camera_settings = {**default_camera_settings}
    for key in default_camera_settings:
        if key in data:
            camera_settings[key] = data[key]

    detection_features = {**default_detection_features}
    if 'detection_features' in data and isinstance(data['detection_features'], dict):
        for key in default_detection_features:
            if key in data['detection_features']:
                detection_features[key] = bool(data['detection_features'][key])

    alert_settings = {**default_alert_settings}
    if 'alert_settings' in data and isinstance(data['alert_settings'], dict):
        for key in default_alert_settings:
            if key in data['alert_settings']:
                alert_settings[key] = data['alert_settings'][key]

    # Load config to get organization_id
    try:
        with open('config.json', 'r') as f:
            config_data = json.load(f)
        organization_id = 17  # Default from config.json
        if 'organizations' in config_data and len(config_data['organizations']) > 0:
            organization_id = config_data['organizations'][0]['id']
    except Exception as e:
        print(f"⚠️ Error loading config.json: {e}")
        organization_id = 17  # Default fallback

    # Update video_bitrate default to match external API
    if camera_settings['video_bitrate'] == 4000:
        camera_settings['video_bitrate'] = 2000

    # Create initial camera data (will be replaced by external API response)
    new_camera = {
        'name': name,
        'location': location or 'Unassigned',
        'ip_address': ip_address,
        'mac_address': data.get('mac_address', ''),
        'model': data.get('model', ''),
        'serial_number': data.get('serial_number', ''),
        'organization_id': organization_id,
        'status': status,
        'username': data.get('username', 'root'),
        'password': data.get('password', ''),
        'port': data.get('port', '554'),
        'path': data.get('path', '/axis-media/media.amp'),
        'url': url_value,
        'URL': url_value,
        'guid': guid,
        'description': data.get('description', name),
        'is_recording': is_recording,
        # Direct camera settings fields
        'frame_rate': camera_settings['frame_rate'],
        'resolution': camera_settings['resolution'],
        'video_bitrate': camera_settings['video_bitrate'],
        'enable_audio': camera_settings['enable_audio'],
        'audio_bitrate': camera_settings['audio_bitrate'],
        'preset': camera_settings['preset'],
        'crf': camera_settings['crf'],
        'segment_time': camera_settings['segment_time'],
        # Structured settings
        'camera_settings': camera_settings,
        'detection_features': detection_features,
        'alert_settings': alert_settings
    }

    # Prepare data for external API
    api_payload = {
        "name": name,
        "url": url_value,
        "location": location or 'Unassigned',
        "ip_address": ip_address,
        "model": data.get('model', ''),
        "status": status,
        "detection_features": detection_features,
        "alert_settings": alert_settings,
        "frame_rate": camera_settings['frame_rate'],
        "resolution": camera_settings['resolution'],
        "video_bitrate": camera_settings['video_bitrate'],
        "enable_audio": camera_settings['enable_audio'],
        "audio_bitrate": camera_settings['audio_bitrate'],
        "preset": camera_settings['preset'],
        "crf": camera_settings['crf'],
        "segment_time": camera_settings['segment_time'],
        "organization_id": organization_id,
        "guid": guid
    }

    # Send to external API
    try:
        print(f"📡 Sending camera data to external API: {name}")
        api_url = f"{get_api_base_url()}/cameras/"
        api_response = requests.post(
            api_url,
            json=api_payload,
            headers={'Content-Type': 'application/json'},
            timeout=10
        )
        
        if api_response.status_code == 200 or api_response.status_code == 201:
            api_data = api_response.json()
            # Replace local camera data entirely with external API response
            new_camera = api_data
            print(f"✅ Camera successfully created in external API: {api_data.get('id', 'unknown')}")
        else:
            print(f"❌ External API error: {api_response.status_code} - {api_response.text}")
            return jsonify({'error': f'External API error: {api_response.status_code}'}), 500
            
    except requests.exceptions.RequestException as e:
        print(f"❌ Failed to connect to external API: {str(e)}")
        return jsonify({'error': f'Failed to connect to external API: {str(e)}'}), 500

    cameras_data.append(new_camera)
    if save_cameras():
        # Update mediamtx.yml with new camera
        if url_value:  # Only add to mediamtx if we have an RTSP URL
            print(f"🔧 Adding camera to MediaMTX: {guid} -> {url_value}")
            if update_mediamtx_config(guid, url_value):
                # Restart MediaMTX to apply new configuration
                print(f"🔄 Restarting MediaMTX for new camera: {guid}")
                restart_mediamtx()
            else:
                print(f"❌ Failed to update MediaMTX config for camera: {guid}")
        else:
            print(f"⚠️ No RTSP URL provided for camera: {guid}")
        return jsonify(new_camera), 201
    # rollback on failure
    cameras_data.pop()
    return jsonify({'error': 'Failed to save camera'}), 500

@app.route('/api/cameras/<camera_id>', methods=['PUT'])
def update_camera(camera_id):
    """Update an existing camera and persist in config.json"""
    global cameras_data
    data = request.get_json(force=True) or {}

    # Find camera by external ID (since we're now using external IDs as primary)
    # Convert camera_id to int for comparison with external API IDs
    try:
        numeric_camera_id = int(camera_id)
    except ValueError:
        return jsonify({'error': 'Invalid camera ID format'}), 400
    
    camera = next((c for c in cameras_data if c.get('id') == numeric_camera_id), None)
    if not camera:
        return jsonify({'error': 'Camera not found'}), 404

    # Track previous recording state
    previous_is_recording = bool(camera.get('is_recording', False))

    # Validate ip uniqueness if changed
    if 'ip_address' in data:
        new_ip = (data.get('ip_address') or '').strip()
        if not new_ip:
            return jsonify({'error': 'ip_address cannot be empty'}), 400
        if new_ip != camera.get('ip_address') and any(c.get('ip_address') == new_ip for c in cameras_data):
            return jsonify({'error': 'Camera with this IP already exists'}), 409
        camera['ip_address'] = new_ip

    # Update simple fields
    for field in ['name', 'location', 'model', 'serial_number', 'mac_address', 'organization_id', 'status', 'description', 'username', 'password', 'port', 'path']:
        if field in data and isinstance(data[field], str):
            camera[field] = data[field].strip()

    # URL aliases
    url_changed = False
    if 'URL' in data or 'url' in data:
        new_url = (data.get('URL') or data.get('url') or '').strip()
        old_url = camera.get('url', '')
        camera['URL'] = new_url
        camera['url'] = new_url
        url_changed = (new_url != old_url)

    # is_recording
    if 'is_recording' in data:
        camera['is_recording'] = bool(data['is_recording'])

    # Update camera settings
    camera_settings_fields = ['frame_rate', 'resolution', 'video_bitrate', 'enable_audio', 
                             'audio_bitrate', 'preset', 'crf', 'segment_time']
    for field in camera_settings_fields:
        if field in data:
            camera[field] = data[field]
            # Also update in camera_settings if it exists
            if 'camera_settings' not in camera:
                camera['camera_settings'] = {}
            camera['camera_settings'][field] = data[field]

    # Update detection features
    if 'detection_features' in data and isinstance(data['detection_features'], dict):
        if 'detection_features' not in camera:
            camera['detection_features'] = {}
        for key, value in data['detection_features'].items():
            camera['detection_features'][key] = bool(value)

    # Update alert settings
    if 'alert_settings' in data and isinstance(data['alert_settings'], dict):
        if 'alert_settings' not in camera:
            camera['alert_settings'] = {}
        for key, value in data['alert_settings'].items():
            camera['alert_settings'][key] = value

    # Update timestamp
    camera['updated_at'] = datetime.now().isoformat()

    # Manage threads if recording state changed
    if bool(camera.get('is_recording', False)) != previous_is_recording:
        if camera.get('is_recording', False):
            start_camera_thread(camera)
        else:
            stop_camera_thread(numeric_camera_id)

    if save_cameras():
        # Call external API for camera update
        try:
            # Prepare data for external API
            api_payload = {
                "name": camera['name'],
                "url": camera.get('url', ''),
                "location": camera.get('location', 'Unassigned'),
                "ip_address": camera['ip_address'],
                "model": camera.get('model', ''),
                "status": camera.get('status', 'Active'),
                "detection_features": camera.get('detection_features', {}),
                "alert_settings": camera.get('alert_settings', {}),
                "frame_rate": camera.get('frame_rate', 30),
                "resolution": camera.get('resolution', '1920x1080'),
                "video_bitrate": camera.get('video_bitrate', 2000),
                "enable_audio": camera.get('enable_audio', True),
                "audio_bitrate": camera.get('audio_bitrate', 128),
                "preset": camera.get('preset', 'medium'),
                "crf": camera.get('crf', 23),
                "segment_time": camera.get('segment_time', 60),
                "organization_id": 17,
                "guid": camera.get('GUID', '')
            }
            
            # Check if camera has external ID (merged into main camera data)
            # Look for external_id field or check if id is numeric (from external API)
            external_camera_id = camera.get('external_id')
            if not external_camera_id:
                # Check if the main ID is numeric (indicating it came from external API)
                main_id = camera.get('id')
                if isinstance(main_id, int):
                    external_camera_id = main_id
            
            if external_camera_id:
                # Update existing camera in external API (handle both string and numeric IDs)
                api_url = f"{get_api_base_url()}/cameras/{external_camera_id}"
                api_response = requests.put(
                    api_url,
                    json=api_payload,
                    headers={'Content-Type': 'application/json'},
                    timeout=10
                )
                
                if api_response.status_code == 200:
                    # Replace local camera data entirely with external response
                    updated_external_data = api_response.json()
                    # Clear existing camera data and replace with external data
                    camera.clear()
                    camera.update(updated_external_data)
                    save_cameras()  # Save the external data
                    print(f"✅ Camera successfully updated in external API: {external_camera_id} (type: {type(external_camera_id)})")
                else:
                    print(f"❌ External API update error: {api_response.status_code} - {api_response.text}")
            else:
                # Create new camera in external API if no external ID exists
                api_url = f"{get_api_base_url()}/cameras/"
                api_response = requests.post(
                    api_url,
                    json=api_payload,
                    headers={'Content-Type': 'application/json'},
                    timeout=10
                )
                
                if api_response.status_code == 200 or api_response.status_code == 201:
                    api_data = api_response.json()
                    external_camera_id = api_data.get('id')
                    if external_camera_id:
                        # Replace local camera data entirely with external data
                        camera.clear()
                        camera.update(api_data)
                        save_cameras()  # Save the external data
                        print(f"✅ Camera successfully created in external API: {external_camera_id} (type: {type(external_camera_id)})")
                else:
                    print(f"❌ External API create error: {api_response.status_code} - {api_response.text}")
                    
        except requests.exceptions.RequestException as e:
            print(f"❌ Failed to connect to external API: {str(e)}")
        
        # Update mediamtx.yml if URL changed
        if url_changed and camera.get('GUID'):
            if camera.get('url'):
                if update_mediamtx_camera(camera['GUID'], camera['url']):
                    # Restart MediaMTX to apply new configuration
                    restart_mediamtx()
            else:
                if remove_mediamtx_config(camera['GUID']):
                    # Restart MediaMTX to apply new configuration
                    restart_mediamtx()
        return jsonify(camera)
    return jsonify({'error': 'Failed to save camera'}), 500

@app.route('/api/merge-camera-external-data', methods=['POST'])
def merge_camera_external_data():
    """Merge external API response data with local camera data in config.json"""
    try:
        data = request.get_json(force=True) or {}
        local_camera_id = data.get('local_camera_id')
        external_camera_data = data.get('external_camera_data')
        
        if not local_camera_id or not external_camera_data:
            return jsonify({'error': 'local_camera_id and external_camera_data are required'}), 400
        
        # Load current config
        with open('config.json', 'r') as f:
            config_data = json.load(f)
        
        # Find and replace the camera with external data
        camera_updated = False
        # Convert local_camera_id to int for comparison with external API IDs
        try:
            numeric_local_camera_id = int(local_camera_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid local camera ID format'}), 400
            
        for i, camera in enumerate(config_data.get('cameras', [])):
            if camera.get('id') == numeric_local_camera_id:
                # Replace local camera data entirely with external data
                config_data['cameras'][i] = external_camera_data
                camera_updated = True
                break
        
        if not camera_updated:
            return jsonify({'error': 'Camera not found'}), 404
        
        # Save updated config
        with open('config.json', 'w') as f:
            json.dump(config_data, f, indent=2)
        
        external_id = external_camera_data.get('id', 'unknown')
        print(f"✅ Merged external data with camera {local_camera_id}: {external_id} (type: {type(external_id)})")
        return jsonify({'success': True, 'merged_data': external_camera_data})
        
    except Exception as e:
        print(f"❌ Error updating camera external data: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cameras/<camera_id>', methods=['DELETE'])
def delete_camera(camera_id):
    """Delete a camera, stop recording thread, and remove from groups."""
    global cameras_data, groups_data
    
    # Convert camera_id to int for comparison with external API IDs
    try:
        numeric_camera_id = int(camera_id)
    except ValueError:
        return jsonify({'error': 'Invalid camera ID format'}), 400
    
    index = next((i for i, c in enumerate(cameras_data) if c.get('id') == numeric_camera_id), None)
    if index is None:
        return jsonify({'error': 'Camera not found'}), 404

    # Get camera GUID before removing
    camera_guid = cameras_data[index].get('guid', cameras_data[index].get('GUID', ''))

    # Stop any active recording
    stop_camera_thread(numeric_camera_id)

    # Remove from cameras
    cameras_data.pop(index)

    # Remove camera id from any groups
    for g in groups_data:
        ids = g.get('camera_ids', [])
        if camera_id in ids:
            g['camera_ids'] = [cid for cid in ids if cid != camera_id]

    if save_cameras():
        # Remove from mediamtx.yml
        if camera_guid:
            if remove_mediamtx_config(camera_guid):
                # Restart MediaMTX to apply new configuration
                restart_mediamtx()
        return jsonify({'success': True})
    return jsonify({'error': 'Failed to save changes'}), 500

@app.route('/api/discover')
def discover_cameras_api():
    """API endpoint to discover devices on the LAN and flag likely cameras."""
    devices = find_lan_devices()
    discovered = []
    for ip_address, mac, vendor in devices:
        vendor_lower = vendor.lower()
        is_camera = any(v in vendor_lower for v in KNOWN_CAMERA_VENDORS)
        discovered.append({
            'ip_address': ip_address,
            'mac_address': mac,
            'vendor': vendor,
            'is_camera': is_camera,
        })
    return jsonify(discovered)
@app.route('/api/groups', methods=['GET'])
def get_groups():
    """List all groups"""
    return jsonify(groups_data)

@app.route('/api/groups', methods=['POST'])
def create_group():
    """Create a new group: expects {name, camera_ids[]}"""
    global groups_data
    data = request.get_json(force=True) or {}
    name = data.get('name', '').strip()
    camera_ids = data.get('camera_ids', [])
    if not name:
        return jsonify({'error': 'Group name is required'}), 400
    # Validate camera IDs
    valid_ids = {c['id'] for c in cameras_data}
    invalid = [cid for cid in camera_ids if cid not in valid_ids]
    if invalid:
        return jsonify({'error': 'Invalid camera ids', 'invalid_ids': invalid}), 400
    group_id = f"grp_{uuid.uuid4().hex[:8]}"
    group = {'id': group_id, 'name': name, 'camera_ids': camera_ids}
    groups_data.append(group)
    if save_cameras():
        return jsonify(group), 201
    return jsonify({'error': 'Failed to save group'}), 500

@app.route('/api/groups/<group_id>', methods=['PUT'])
def update_group(group_id):
    """Update group name and/or camera_ids"""
    global groups_data
    data = request.get_json(force=True) or {}
    # Find group
    group = next((g for g in groups_data if g.get('id') == group_id), None)
    if not group:
        return jsonify({'error': 'Group not found'}), 404
    # Update fields
    if 'name' in data and isinstance(data['name'], str):
        group['name'] = data['name'].strip()
    if 'camera_ids' in data and isinstance(data['camera_ids'], list):
        valid_ids = {c['id'] for c in cameras_data}
        invalid = [cid for cid in data['camera_ids'] if cid not in valid_ids]
        if invalid:
            return jsonify({'error': 'Invalid camera ids', 'invalid_ids': invalid}), 400
        group['camera_ids'] = data['camera_ids']
    if save_cameras():
        return jsonify(group)
    return jsonify({'error': 'Failed to save group'}), 500

@app.route('/api/groups/<group_id>', methods=['DELETE'])
def delete_group(group_id):
    """Delete a group"""
    global groups_data
    index = next((i for i, g in enumerate(groups_data) if g.get('id') == group_id), None)
    if index is None:
        return jsonify({'error': 'Group not found'}), 404
    groups_data.pop(index)
    if save_cameras():
        return jsonify({'success': True})
    return jsonify({'error': 'Failed to delete group'}), 500

@app.route('/api/discover/cameras')
def discover_only_cameras_api():
    """API endpoint to discover and return only likely camera devices."""
    devices = find_lan_devices()
    cameras_only = []
    for ip_address, mac, vendor in devices:
        vendor_lower = vendor.lower()
        if any(v in vendor_lower for v in KNOWN_CAMERA_VENDORS):
            cameras_only.append({
                'ip_address': ip_address,
                'mac_address': mac,
                'vendor': vendor,
            })
    return jsonify(cameras_only)

@app.route('/api/cameras/<camera_id>/toggle', methods=['POST'])
def toggle_camera_recording(camera_id):
    """API endpoint to toggle camera recording status"""
    global cameras_data
    
    # Convert camera_id to int for comparison with external API IDs
    try:
        numeric_camera_id = int(camera_id)
    except ValueError:
        return jsonify({'error': 'Invalid camera ID format'}), 400
    
    # Find the camera
    camera = None
    for cam in cameras_data:
        if cam['id'] == numeric_camera_id:
            camera = cam
            break
    
    if not camera:
        return jsonify({'error': 'Camera not found'}), 404
    
    # Get the new recording status
    data = request.get_json()
    new_recording_status = data.get('is_recording', False)
    
    # Update the recording status
    camera['is_recording'] = new_recording_status
    
    # Manage thread based on recording status
    if new_recording_status:
        # Start thread if recording is turned ON
        start_camera_thread(camera)
        print(f"Recording turned ON for camera: {camera['name']}")
    else:
        # Stop thread if recording is turned OFF
        stop_camera_thread(numeric_camera_id)
        print(f"Recording turned OFF for camera: {camera['name']}")
    
    # Save to file
    if save_cameras():
        return jsonify({'success': True, 'camera': camera})
    else:
        return jsonify({'error': 'Failed to save changes'}), 500

@app.route('/api/threads/status')
def get_thread_status():
    """API endpoint to get current thread status"""
    thread_status = {}
    for camera in cameras_data:
        camera_id = camera['id']
        thread_status[camera_id] = {
            'is_recording': camera.get('is_recording', False),
            'has_active_thread': camera_id in active_threads,
            'camera_name': camera['name']
        }
    return jsonify(thread_status)

@app.route('/api/video/upload/status')
def get_upload_status():
    """API endpoint to get video upload configuration and status"""
    return jsonify({
        'upload_enabled': video_config['upload_to_api'],
        'api_url': video_config['api_url'],
        'retry_attempts': video_config.get('upload_retry_attempts', 3),
        'timeout': video_config.get('upload_timeout', 30),
        'delay_between_files': video_config.get('upload_delay_between_files', 0.5),
        'organization_id': api_organization_id,
        'store_locally': video_config['store_locally'],
        'dedicated_upload_thread_running': upload_thread_running,
        'uploaded_files_count': len(uploaded_files)
    })

@app.route('/api/video/upload/thread/start', methods=['POST'])
def start_upload_thread():
    """Start the dedicated upload thread"""
    try:
        start_dedicated_upload_thread()
        return jsonify({'message': 'Dedicated upload thread started successfully'})
    except Exception as e:
        return jsonify({'error': f'Failed to start upload thread: {str(e)}'}), 500

@app.route('/api/video/upload/thread/stop', methods=['POST'])
def stop_upload_thread():
    """Stop the dedicated upload thread"""
    try:
        stop_dedicated_upload_thread()
        return jsonify({'message': 'Dedicated upload thread stopped successfully'})
    except Exception as e:
        return jsonify({'error': f'Failed to stop upload thread: {str(e)}'}), 500

@app.route('/api/video/upload/thread/status')
def get_upload_thread_status():
    """Get the status of the dedicated upload thread"""
    return jsonify({
        'running': upload_thread_running,
        'uploaded_files_count': len(uploaded_files),
        'uploaded_files': list(uploaded_files)[-10:] if uploaded_files else []  # Last 10 uploaded files
    })

@app.route('/api/cameras/<camera_id>/frame')
def get_camera_frame(camera_id):
    """API endpoint to get camera frame as MJPEG stream"""
    # Convert camera_id to int for comparison with external API IDs
    try:
        numeric_camera_id = int(camera_id)
    except ValueError:
        return jsonify({'error': 'Invalid camera ID format'}), 400
    
    # Find the camera
    camera = None
    for cam in cameras_data:
        if cam['id'] == numeric_camera_id:
            camera = cam
            break
    
    if not camera:
        return jsonify({'error': 'Camera not found'}), 404
    
    # Get RTSP URL from query parameter or camera config
    rtsp_url = request.args.get('rtsp_url')
    if not rtsp_url:
        rtsp_url = camera.get('URL', '') or camera.get('url', '')
    
    if not rtsp_url:
        return jsonify({'error': 'Camera URL not configured'}), 400
    
    try:
        # Use ffmpeg to convert RTSP stream to MJPEG
        ffmpeg_cmd = [
            "ffmpeg",
            "-rtsp_transport", "tcp",
            "-i", rtsp_url,
            "-f", "mjpeg",
            "-q:v", "5",  # Quality level (1-31, lower is better)
            "-r", "10",   # Frame rate
            "-s", "640x360",  # Resolution for live view
            "pipe:1"
        ]
        
        process = subprocess.Popen(
            ffmpeg_cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE
        )
        
        def generate():
            try:
                while True:
                    # Read a chunk of data
                    chunk = process.stdout.read(1024)
                    if not chunk:
                        break
                    yield chunk
            except Exception as e:
                print(f"Error in MJPEG stream: {e}")
            finally:
                if process.poll() is None:
                    process.terminate()
        
        return app.response_class(
            generate(),
            mimetype='multipart/x-mixed-replace; boundary=frame'
        )
        
    except Exception as e:
        print(f"Error creating MJPEG stream: {e}")
        return jsonify({'error': 'Failed to create video stream'}), 500


@app.route('/api/cameras/<camera_id>/hls/<path:filename>')
def get_hls_file(camera_id, filename):
    """Serve HLS files for live streaming"""
    stream_dir = os.path.join(base_video_path, 'live_streams', camera_id)
    return send_from_directory(stream_dir, filename)


@app.route('/api/videos')
def get_videos():
    """API endpoint to get list of videos by date and optional camera ID and hour"""
    date = request.args.get('date')
    camera_id = request.args.get('camera_id')
    hour = request.args.get('hour')
    
    if not date:
        return jsonify({'error': 'Date parameter is required'}), 400
    
    try:
        videos = []
        date_path = os.path.join(base_video_path, date)
        
        if not os.path.exists(date_path):
            return jsonify(videos)
        
        # Get all camera GUIDs for the date
        for item in os.listdir(date_path):
            item_path = os.path.join(date_path, item)
            if os.path.isdir(item_path):
                # Find camera by GUID
                camera = None
                for cam in cameras_data:
                    if cam.get('GUID') == item:
                        camera = cam
                        break
                
                # Skip if camera_id filter is specified and doesn't match
                if camera_id and camera and camera.get('id') != camera_id:
                    continue
                
                # Look for hour folders (00, 01, 02, ..., 23)
                for hour_folder in os.listdir(item_path):
                    hour_path = os.path.join(item_path, hour_folder)
                    if os.path.isdir(hour_path) and hour_folder.isdigit() and 0 <= int(hour_folder) <= 23:
                        # Skip if hour filter is specified and doesn't match
                        if hour and hour_folder != hour:
                            continue
                            
                        # Get all video files in this hour folder
                        for video_file in os.listdir(hour_path):
                            if video_file.endswith('.mp4'):
                                video_path = os.path.join(hour_path, video_file)
                                file_stat = os.stat(video_path)
                                
                                video_info = {
                                    'filename': video_file,
                                    'path': video_path,
                                    'camera_id': camera.get('id') if camera else None,
                                    'camera_name': camera.get('name') if camera else 'Unknown Camera',
                                    'camera_guid': item,
                                    'hour': hour_folder,
                                    'size': file_stat.st_size,
                                    'created_at': datetime.fromtimestamp(file_stat.st_ctime).isoformat(),
                                    'modified_at': datetime.fromtimestamp(file_stat.st_mtime).isoformat()
                                }
                                videos.append(video_info)
        
        # Sort by creation time (newest first)
        videos.sort(key=lambda x: x['created_at'], reverse=True)
        return jsonify(videos)
        
    except Exception as e:
        print(f"Error getting videos: {e}")
        return jsonify({'error': 'Failed to get videos'}), 500

@app.route('/api/videos/play/<filename>')
def play_video(filename):
    """API endpoint to stream video for playback"""
    try:
        # Find the video file
        video_path = None
        for root, dirs, files in os.walk(base_video_path):
            if filename in files:
                video_path = os.path.join(root, filename)
                break
        
        if not video_path or not os.path.exists(video_path):
            return jsonify({'error': 'Video not found'}), 404
        
        return send_file(video_path, mimetype='video/mp4')
        
    except Exception as e:
        print(f"Error playing video: {e}")
        return jsonify({'error': 'Failed to play video'}), 500

@app.route('/api/videos/download/<filename>')
def download_video(filename):
    """API endpoint to download video file"""
    try:
        # Find the video file
        video_path = None
        for root, dirs, files in os.walk(base_video_path):
            if filename in files:
                video_path = os.path.join(root, filename)
                break
        
        if not video_path or not os.path.exists(video_path):
            return jsonify({'error': 'Video not found'}), 404
        
        return send_file(video_path, as_attachment=True, download_name=filename)
        
    except Exception as e:
        print(f"Error downloading video: {e}")
        return jsonify({'error': 'Failed to download video'}), 500

# Authentication Routes
@app.route('/login')
def login_page():
    """Serve the login page"""
    return send_file('login.html')

@app.route('/api/auth/login', methods=['POST'])
def login():
    """Handle user login"""
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({'message': 'Username and password are required'}), 400
        
        # Simple authentication - in production, use proper password hashing
        # For demo purposes, accept any username/password or admin/admin123
        if (username == 'admin' and password == 'admin123') or (username and password):
            # Create user session
            user = {
                'id': f"user_{str(uuid.uuid4())[:8]}",
                'username': username,
                'name': username.title(),
                'role': 'Admin' if username == 'admin' else 'User',
                'organization_id': 'org_001',
                'isfactory': True,
                'login_time': datetime.now().isoformat()
            }
            
            session['user'] = user
            session['is_authenticated'] = True
            
            # Fetch cameras from API after successful login
            try:
                fetch_cameras_from_api()
            except Exception as e:
                print(f"Warning: Failed to fetch cameras after login: {e}")
            
            return jsonify({
                'message': 'Login successful',
                'user': user
            })
        else:
            return jsonify({'message': 'Invalid username or password'}), 401
            
    except Exception as e:
        print(f"Login error: {e}")
        return jsonify({'message': 'Login failed'}), 500

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    """Handle user logout"""
    try:
        session.clear()
        return jsonify({'message': 'Logout successful'})
    except Exception as e:
        print(f"Logout error: {e}")
        return jsonify({'message': 'Logout failed'}), 500

@app.route('/api/auth/status')
def auth_status():
    """Check authentication status"""
    if session.get('is_authenticated'):
        return jsonify({
            'authenticated': True,
            'user': session.get('user')
        })
    return jsonify({'authenticated': False})

@app.route('/signup')
def signup_page():
    """Serve the signup page"""
    return send_file('signup.html')

@app.route('/api/auth/signup', methods=['POST'])
def signup():
    """Handle user signup with organization creation - simplified API"""
    try:
        data = request.get_json()
        name = data.get('name')
        email = data.get('email')
        password = data.get('password')
        organization_name = data.get('organization_name')
        organization_description = data.get('organization_description', '')
        
        if not name or not email or not password or not organization_name:
            return jsonify({'message': 'Name, email, password, and organization name are required'}), 400
        
        # Load current config
        with open('config.json', 'r') as f:
            config = json.load(f)
        
        # Check if user already exists
        users = config.get('users', [])
        if any(user.get('email') == email for user in users):
            return jsonify({'message': 'Email already exists'}), 400
        
        # Create new organization
        new_org = {
            'id': f"org_{str(uuid.uuid4())[:8]}",
            'name': organization_name,
            'description': organization_description,
            'address': '',
            'contact_email': email,
            'retention': 30,
            'upload_username': 'admin',
            'upload_password': 'admin123',
            'storage_provider': '',
            'server_video_path': '',
            'client_video_access_path': '',
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat()
        }
        
        # Create new user
        new_user = {
            'id': f"user_{str(uuid.uuid4())[:8]}",
            'name': name,
            'email': email,
            'password': password,  # In production, hash this password
            'role': 'Admin',
            'login_role': 'Owner',
            'organization_id': new_org['id'],
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat()
        }
        
        # Initialize arrays if they don't exist
        if 'organizations' not in config:
            config['organizations'] = []
        if 'users' not in config:
            config['users'] = []
        
        # Add new organization and user
        config['organizations'].append(new_org)
        config['users'].append(new_user)
        
        # Save config
        with open('config.json', 'w') as f:
            json.dump(config, f, indent=2)
        
        return jsonify({
            'message': 'Account created successfully',
            'user': {
                'id': new_user['id'],
                'name': new_user['name'],
                'email': new_user['email'],
                'role': new_user['role'],
                'organization_id': new_user['organization_id']
            },
            'organization': {
                'id': new_org['id'],
                'name': new_org['name'],
                'description': new_org['description']
            }
        }), 201
        
    except Exception as e:
        print(f"Signup error: {e}")
        return jsonify({'message': 'Failed to create account'}), 500

# Static file routes
@app.route('/<path:filename>')
def serve_static(filename):
    """Serve static files like images, CSS, JS"""
    try:
        return send_file(filename)
    except FileNotFoundError:
        return "File not found", 404

# Organization Management Routes
@app.route('/organizations')
def organizations_page():
    """Serve the organizations page"""
    return send_file('organizations.html')

@app.route('/api/organizations', methods=['GET'])
def get_organizations():
    """Get all organizations"""
    try:
        with open('config.json', 'r') as f:
            config = json.load(f)
        
        # Return organizations array if it exists, otherwise return current organization as single item
        if 'organizations' in config:
            return jsonify(config['organizations'])
        else:
            # Convert single organization to array format
            org = config.get('organization', {})
            if org:
                return jsonify([org])
            return jsonify([])
    except Exception as e:
        print(f"Error loading organizations: {e}")
        return jsonify({'error': 'Failed to load organizations'}), 500

@app.route('/api/organizations', methods=['POST'])
def create_organization():
    """Create a new organization"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['name', 'contact_email', 'address']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400
        
        # Load current config
        with open('config.json', 'r') as f:
            config = json.load(f)
        
        # Create new organization
        new_org = {
            'id': f"org_{str(uuid.uuid4())[:8]}",
            'name': data['name'],
            'address': data['address'],
            'contact_email': data['contact_email'],
            'storage_provider': data.get('storage_provider', ''),
            'server_video_path': data.get('server_video_path', ''),
            'client_video_access_path': data.get('client_video_access_path', ''),
            'retention': data.get('retention', 10),
            'upload_username': data.get('upload_username', 'admin'),
            'upload_password': data.get('upload_password', 'admin'),
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat()
        }
        
        # Initialize organizations array if it doesn't exist
        if 'organizations' not in config:
            config['organizations'] = []
        
        # Add new organization
        config['organizations'].append(new_org)
        
        # Save config
        with open('config.json', 'w') as f:
            json.dump(config, f, indent=2)
        
        return jsonify(new_org), 201
        
    except Exception as e:
        print(f"Error creating organization: {e}")
        return jsonify({'error': 'Failed to create organization'}), 500

@app.route('/api/organizations/<org_id>', methods=['PUT'])
def update_organization(org_id):
    """Update an existing organization"""
    try:
        data = request.get_json()
        
        # Load current config
        with open('config.json', 'r') as f:
            config = json.load(f)
        
        # Find organization to update
        organizations = config.get('organizations', [])
        org_index = None
        
        for i, org in enumerate(organizations):
            if org['id'] == org_id:
                org_index = i
                break
        
        if org_index is None:
            return jsonify({'error': 'Organization not found'}), 404
        
        # Update organization
        organizations[org_index].update({
            'name': data.get('name', organizations[org_index]['name']),
            'address': data.get('address', organizations[org_index]['address']),
            'contact_email': data.get('contact_email', organizations[org_index]['contact_email']),
            'storage_provider': data.get('storage_provider', organizations[org_index].get('storage_provider', '')),
            'server_video_path': data.get('server_video_path', organizations[org_index].get('server_video_path', '')),
            'client_video_access_path': data.get('client_video_access_path', organizations[org_index].get('client_video_access_path', '')),
            'retention': data.get('retention', organizations[org_index].get('retention', 10)),
            'upload_username': data.get('upload_username', organizations[org_index].get('upload_username', 'admin')),
            'upload_password': data.get('upload_password', organizations[org_index].get('upload_password', 'admin')),
            'updated_at': datetime.now().isoformat()
        })
        
        # Save config
        with open('config.json', 'w') as f:
            json.dump(config, f, indent=2)
        
        return jsonify(organizations[org_index])
        
    except Exception as e:
        print(f"Error updating organization: {e}")
        return jsonify({'error': 'Failed to update organization'}), 500

@app.route('/api/organizations/<org_id>', methods=['DELETE'])
def delete_organization(org_id):
    """Delete an organization"""
    try:
        # Load current config
        with open('config.json', 'r') as f:
            config = json.load(f)
        
        # Find organization to delete
        organizations = config.get('organizations', [])
        org_index = None
        
        for i, org in enumerate(organizations):
            if org['id'] == org_id:
                org_index = i
                break
        
        if org_index is None:
            return jsonify({'error': 'Organization not found'}), 404
        
        # Remove organization
        deleted_org = organizations.pop(org_index)
        
        # Save config
        with open('config.json', 'w') as f:
            json.dump(config, f, indent=2)
        
        return jsonify({'message': 'Organization deleted successfully'})
        
    except Exception as e:
        print(f"Error deleting organization: {e}")
        return jsonify({'error': 'Failed to delete organization'}), 500

@app.route('/api/organizations/export', methods=['GET'])
def export_organizations():
    """Export organizations data in various formats"""
    try:
        format_type = request.args.get('format', 'json').lower()
        
        # Load current config
        with open('config.json', 'r') as f:
            config = json.load(f)
        
        organizations = config.get('organizations', [])
        
        if format_type == 'csv':
            import csv
            import io
            
            output = io.StringIO()
            if organizations:
                fieldnames = organizations[0].keys()
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(organizations)
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'text/csv'
            response.headers['Content-Disposition'] = 'attachment; filename=organizations.csv'
            return response
            
        elif format_type == 'xlsx':
            try:
                import openpyxl
                from openpyxl import Workbook
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Organizations"
                
                if organizations:
                    # Write headers
                    headers = list(organizations[0].keys())
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                    
                    # Write data
                    for row, org in enumerate(organizations, 2):
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=row, column=col, value=org.get(header, ''))
                
                # Save to bytes
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response.headers['Content-Disposition'] = 'attachment; filename=organizations.xlsx'
                return response
                
            except ImportError:
                return jsonify({'error': 'openpyxl library not installed for Excel export'}), 500
                
        else:  # Default to JSON
            response = make_response(json.dumps(organizations, indent=2))
            response.headers['Content-Type'] = 'application/json'
            response.headers['Content-Disposition'] = 'attachment; filename=organizations.json'
            return response
            
    except Exception as e:
        print(f"Export error: {e}")
        return jsonify({'error': 'Failed to export organizations'}), 500

@app.route('/api/organizations/import', methods=['POST'])
def import_organizations():
    """Import organizations data from file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Load current config
        with open('config.json', 'r') as f:
            config = json.load(f)
        
        # Parse file based on extension
        filename = file.filename.lower()
        
        if filename.endswith('.json'):
            import_data = json.load(file)
        elif filename.endswith('.csv'):
            import csv
            import io
            
            content = file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(content))
            import_data = list(csv_reader)
        else:
            return jsonify({'error': 'Unsupported file format. Use JSON or CSV.'}), 400
        
        # Validate and process import data
        if not isinstance(import_data, list):
            return jsonify({'error': 'Invalid data format. Expected array of organizations.'}), 400
        
        # Initialize organizations array if it doesn't exist
        if 'organizations' not in config:
            config['organizations'] = []
        
        imported_count = 0
        for org_data in import_data:
            # Generate new ID if not provided
            if 'id' not in org_data or not org_data['id']:
                org_data['id'] = f"org_{str(uuid.uuid4())[:8]}"
            
            # Add timestamps if not provided
            if 'created_at' not in org_data:
                org_data['created_at'] = datetime.now().isoformat()
            if 'updated_at' not in org_data:
                org_data['updated_at'] = datetime.now().isoformat()
            
            # Check if organization already exists
            existing_org = next((org for org in config['organizations'] if org['id'] == org_data['id']), None)
            if existing_org:
                # Update existing organization
                existing_org.update(org_data)
                existing_org['updated_at'] = datetime.now().isoformat()
            else:
                # Add new organization
                config['organizations'].append(org_data)
                imported_count += 1
        
        # Save config
        with open('config.json', 'w') as f:
            json.dump(config, f, indent=2)
        
        return jsonify({
            'message': f'Successfully imported {imported_count} new organizations',
            'total_organizations': len(config['organizations'])
        })
        
    except Exception as e:
        print(f"Import error: {e}")
        return jsonify({'error': 'Failed to import organizations'}), 500

@app.route('/profile')
def profile_page():
    """Serve the profile page"""
    return send_file('profile.html')

@app.route('/api/users/profile', methods=['GET'])
def get_user_profile():
    """Get current user's profile"""
    try:
        if not session.get('is_authenticated'):
            return jsonify({'error': 'Not authenticated'}), 401
        
        user_id = session.get('user', {}).get('id')
        if not user_id:
            return jsonify({'error': 'User ID not found'}), 400
        
        # Load current config
        with open('config.json', 'r') as f:
            config = json.load(f)
        
        # Find user
        users = config.get('users', [])
        user = next((u for u in users if u.get('id') == user_id), None)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        return jsonify({
            'success': True,
            'user': user
        })
        
    except Exception as e:
        print(f"Get user profile error: {e}")
        return jsonify({'error': 'Failed to get user profile'}), 500

@app.route('/api/users/profile', methods=['PUT'])
def update_user_profile():
    """Update current user's profile"""
    try:
        if not session.get('is_authenticated'):
            return jsonify({'error': 'Not authenticated'}), 401
        
        user_id = session.get('user', {}).get('id')
        if not user_id:
            return jsonify({'error': 'User ID not found'}), 400
        
        data = request.get_json()
        
        # Load current config
        with open('config.json', 'r') as f:
            config = json.load(f)
        
        # Find user
        users = config.get('users', [])
        user_index = next((i for i, u in enumerate(users) if u.get('id') == user_id), None)
        
        if user_index is None:
            return jsonify({'error': 'User not found'}), 404
        
        # Update user data
        user = users[user_index]
        
        # Update basic fields
        if 'name' in data:
            user['name'] = data['name']
        if 'email' in data:
            user['email'] = data['email']
        
        # Handle password change
        if data.get('current_password') and data.get('new_password'):
            if user.get('password') != data['current_password']:
                return jsonify({'error': 'Current password is incorrect'}), 400
            
            user['password'] = data['new_password']
        
        user['updated_at'] = datetime.now().isoformat()
        
        # Save config
        with open('config.json', 'w') as f:
            json.dump(config, f, indent=2)
        
        # Update session
        session['user'] = user
        
        return jsonify({
            'success': True,
            'message': 'Profile updated successfully',
            'user': user
        })
        
    except Exception as e:
        print(f"Update user profile error: {e}")
        return jsonify({'error': 'Failed to update profile'}), 500

@app.route('/api/organizations/profile', methods=['GET'])
def get_organization_profile():
    """Get current user's organization profile"""
    try:
        if not session.get('is_authenticated'):
            return jsonify({'error': 'Not authenticated'}), 401
        
        user_id = session.get('user', {}).get('id')
        organization_id = session.get('user', {}).get('organization_id')
        
        if not user_id or not organization_id:
            return jsonify({'error': 'User or organization ID not found'}), 400
        
        # Load current config
        with open('config.json', 'r') as f:
            config = json.load(f)
        
        # Find organization
        organizations = config.get('organizations', [])
        organization = next((org for org in organizations if org.get('id') == organization_id), None)
        
        if not organization:
            return jsonify({'error': 'Organization not found'}), 404
        
        return jsonify({
            'success': True,
            'organization': organization
        })
        
    except Exception as e:
        print(f"Get organization profile error: {e}")
        return jsonify({'error': 'Failed to get organization profile'}), 500

@app.route('/api/organizations/profile', methods=['PUT'])
def update_organization_profile():
    """Update current user's organization profile"""
    try:
        if not session.get('is_authenticated'):
            return jsonify({'error': 'Not authenticated'}), 401
        
        user_id = session.get('user', {}).get('id')
        organization_id = session.get('user', {}).get('organization_id')
        
        if not user_id or not organization_id:
            return jsonify({'error': 'User or organization ID not found'}), 400
        
        data = request.get_json()
        
        # Load current config
        with open('config.json', 'r') as f:
            config = json.load(f)
        
        # Find organization
        organizations = config.get('organizations', [])
        org_index = next((i for i, org in enumerate(organizations) if org.get('id') == organization_id), None)
        
        if org_index is None:
            return jsonify({'error': 'Organization not found'}), 404
        
        # Update organization data
        organization = organizations[org_index]
        
        # Update all fields
        for key, value in data.items():
            if value is not None and value != '':
                organization[key] = value
        
        organization['updated_at'] = datetime.now().isoformat()
        
        # Save config
        with open('config.json', 'w') as f:
            json.dump(config, f, indent=2)
        
        return jsonify({
            'success': True,
            'message': 'Organization updated successfully',
            'organization': organization
        })
        
    except Exception as e:
        print(f"Update organization profile error: {e}")
        return jsonify({'error': 'Failed to update organization'}), 500

@app.route('/api/config', methods=['GET'])
def get_config():
    """Get current config.json content"""
    try:
        with open('config.json', 'r') as f:
            config = json.load(f)
        return jsonify(config)
    except Exception as e:
        print(f"Error reading config: {e}")
        return jsonify({'error': 'Failed to read config'}), 500

@app.route('/api/config', methods=['PUT'])
def update_config():
    """Update config.json content"""
    try:
        data = request.get_json()
        
        # Validate that we have the required structure
        if not isinstance(data, dict):
            return jsonify({'error': 'Invalid config data'}), 400
        
        # Write updated config to file
        with open('config.json', 'w') as f:
            json.dump(data, f, indent=2)
        
        return jsonify({'message': 'Config updated successfully'})
    except Exception as e:
        print(f"Error updating config: {e}")
        return jsonify({'error': 'Failed to update config'}), 500

@app.route('/api/mediamtx/config', methods=['GET'])
def get_mediamtx_config():
    """Get MediaMTX configuration"""
    try:
        config_path = 'mediamtx/mediamtx/mediamtx.yml'
        if not os.path.exists(config_path):
            return jsonify({'error': 'MediaMTX config file not found'}), 404
        
        # Parse YAML config to extract port information
        import yaml
        with open(config_path, 'r') as f:
            config = yaml.safe_load(f)
        
        # Extract port information
        result = {
            'webrtcPort': config.get('webrtcAddress', ':8889').split(':')[-1],
            'hlsPort': config.get('hlsAddress', ':8888').split(':')[-1],
            'rtspPort': config.get('rtspAddress', ':8554').split(':')[-1],
            'rtmpPort': config.get('rtmpAddress', ':1935').split(':')[-1],
            'apiPort': config.get('apiAddress', ':9997').split(':')[-1],
            'metricsPort': config.get('metricsAddress', ':9998').split(':')[-1]
        }
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/mediamtx/config', methods=['POST'])
def save_mediamtx_config():
    """Save MediaMTX configuration"""
    try:
        data = request.get_json()
        config_path = 'mediamtx/mediamtx/mediamtx.yml'
        
        if not os.path.exists(config_path):
            return jsonify({'error': 'MediaMTX config file not found'}), 404
        
        # Read current config
        import yaml
        with open(config_path, 'r') as f:
            config = yaml.safe_load(f)
        
        # Update port configurations
        if 'webrtcPort' in data:
            config['webrtcAddress'] = f":{data['webrtcPort']}"
        if 'hlsPort' in data:
            config['hlsAddress'] = f":{data['hlsPort']}"
        if 'rtspPort' in data:
            config['rtspAddress'] = f":{data['rtspPort']}"
        if 'rtmpPort' in data:
            config['rtmpAddress'] = f":{data['rtmpPort']}"
        if 'apiPort' in data:
            config['apiAddress'] = f":{data['apiPort']}"
        if 'metricsPort' in data:
            config['metricsAddress'] = f":{data['metricsPort']}"
        
        # Write updated config
        with open(config_path, 'w') as f:
            yaml.dump(config, f, default_flow_style=False, sort_keys=False)
        
        return jsonify({'message': 'MediaMTX configuration updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/mediamtx/status', methods=['GET'])
def get_mediamtx_status():
    """Get MediaMTX service status"""
    try:
        # Check if MediaMTX process is running
        import psutil
        for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
            try:
                if 'mediamtx' in proc.info['name'].lower() or any('mediamtx' in str(cmd).lower() for cmd in proc.info['cmdline']):
                    return jsonify({'running': True, 'pid': proc.info['pid']})
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                continue
        
        return jsonify({'running': False})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/mediamtx/restart', methods=['POST'])
def restart_mediamtx():
    """Restart MediaMTX service"""
    try:
        # Find and kill existing MediaMTX process
        import psutil
        for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
            try:
                if 'mediamtx' in proc.info['name'].lower() or any('mediamtx' in str(cmd).lower() for cmd in proc.info['cmdline']):
                    proc.terminate()
                    proc.wait(timeout=5)
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.TimeoutExpired):
                continue
        
        # Start MediaMTX in a new thread
        def start_mediamtx():
            time.sleep(2)  # Wait a moment before restarting
            start_mediamtx_server()
        
        threading.Thread(target=start_mediamtx, daemon=True).start()
        
        return jsonify({'message': 'MediaMTX restart initiated'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/mediamtx/save', methods=['POST'])
def save_mediamtx_yaml():
    """Save MediaMTX YAML configuration file"""
    try:
        config_path = 'mediamtx/mediamtx/mediamtx.yml'
        
        # Get the YAML content from request body
        yaml_content = request.get_data(as_text=True)
        
        if not yaml_content:
            return jsonify({'error': 'No YAML content provided'}), 400
        
        # Write the updated YAML content to file
        with open(config_path, 'w') as f:
            f.write(yaml_content)
        
        return jsonify({'message': 'MediaMTX YAML configuration saved successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/mediamtx/mediamtx/mediamtxconfig')
def serve_mediamtx_yaml():
    """Serve MediaMTX YAML configuration file"""
    try:
        config_path = 'mediamtx/mediamtx/mediamtx.yml'
        return send_file(config_path, mimetype='text/yaml')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def update_mediamtx_config(camera_guid, rtsp_url):
    """Update mediamtx.yml file with new camera entry using YAML"""
    try:
        import yaml
        mediamtx_file = 'mediamtx/mediamtx/mediamtx.yml'
        
        # Read the current configuration
        with open(mediamtx_file, 'r', encoding='utf-8') as file:
            config = yaml.safe_load(file)
        
        # Ensure paths section exists
        if 'paths' not in config:
            config['paths'] = {}
        
        # Check if camera already exists
        if camera_guid in config['paths']:
            print(f"Camera {camera_guid} already exists in mediamtx.yml")
            return True
        
        # Add the new path
        config['paths'][camera_guid] = {
            'source': rtsp_url
        }
        
        # Write back to file
        with open(mediamtx_file, 'w', encoding='utf-8') as file:
            yaml.dump(config, file, default_flow_style=False, allow_unicode=True, sort_keys=False)
        
        print(f"✅ Added camera {camera_guid} to mediamtx.yml with source: {rtsp_url}")
        return True
        
    except Exception as e:
        print(f"❌ Error updating mediamtx.yml: {e}")
        return False

def remove_mediamtx_config(camera_guid):
    """Remove camera entry from mediamtx.yml file using YAML"""
    try:
        import yaml
        mediamtx_file = 'mediamtx/mediamtx/mediamtx.yml'
        
        # Read the current configuration
        with open(mediamtx_file, 'r', encoding='utf-8') as file:
            config = yaml.safe_load(file)
        
        # Check if paths section exists and camera exists
        if 'paths' not in config or camera_guid not in config['paths']:
            print(f"Camera {camera_guid} not found in mediamtx.yml")
            return True
        
        # Remove the camera entry
        del config['paths'][camera_guid]
        
        # Write back to file
        with open(mediamtx_file, 'w', encoding='utf-8') as file:
            yaml.dump(config, file, default_flow_style=False, allow_unicode=True, sort_keys=False)
        
        print(f"✅ Removed camera {camera_guid} from mediamtx.yml")
        return True
        
    except Exception as e:
        print(f"❌ Error removing camera from mediamtx.yml: {e}")
        return False

def update_mediamtx_camera(camera_guid, rtsp_url):
    """Update existing camera entry in mediamtx.yml file using YAML"""
    try:
        import yaml
        mediamtx_file = 'mediamtx/mediamtx/mediamtx.yml'
        
        # Read the current configuration
        with open(mediamtx_file, 'r', encoding='utf-8') as file:
            config = yaml.safe_load(file)
        
        # Ensure paths section exists
        if 'paths' not in config:
            config['paths'] = {}
        
        # Update the camera entry
        config['paths'][camera_guid] = {
            'source': rtsp_url
        }
        
        # Write back to file
        with open(mediamtx_file, 'w', encoding='utf-8') as file:
            yaml.dump(config, file, default_flow_style=False, allow_unicode=True, sort_keys=False)
        
        print(f"✅ Updated camera {camera_guid} in mediamtx.yml with source: {rtsp_url}")
        return True
        
    except Exception as e:
        print(f"❌ Error updating camera in mediamtx.yml: {e}")
        return False

@app.route('/api/mediamtx/status', methods=['GET'])
def mediamtx_status_api():
    """Get MediaMTX server status"""
    status = get_mediamtx_status()
    return jsonify(status)

@app.route('/api/mediamtx/start', methods=['POST'])
def mediamtx_start_api():
    """Start MediaMTX server"""
    success = start_mediamtx()
    status = get_mediamtx_status()
    return jsonify({
        'success': success,
        'status': status,
        'message': 'MediaMTX server started' if success else 'Failed to start MediaMTX server'
    })

@app.route('/api/mediamtx/stop', methods=['POST'])
def mediamtx_stop_api():
    """Stop MediaMTX server"""
    success = stop_mediamtx()
    status = get_mediamtx_status()
    return jsonify({
        'success': success,
        'status': status,
        'message': 'MediaMTX server stopped' if success else 'Failed to stop MediaMTX server'
    })

@app.route('/api/mediamtx/restart', methods=['POST'])
def mediamtx_restart_api():
    """Restart MediaMTX server"""
    success = restart_mediamtx()
    status = get_mediamtx_status()
    return jsonify({
        'success': success,
        'status': status,
        'message': 'MediaMTX server restarted' if success else 'Failed to restart MediaMTX server'
    })

def start_mediamtx():
    """Start the MediaMTX server process with visible console"""
    global mediamtx_process
    try:
        if mediamtx_process is None or mediamtx_process.poll() is not None:
            # Check if exe file exists
            if not os.path.exists(mediamtx_exe_path):
                print(f"❌ MediaMTX executable not found: {mediamtx_exe_path}")
                return False
            
            # Check if config file exists
            if not os.path.exists(mediamtx_config_file):
                print(f"❌ MediaMTX config file not found: {mediamtx_config_file}")
                return False
            
            # Start MediaMTX process with CREATE_NEW_CONSOLE flag (Windows)
            mediamtx_process = subprocess.Popen(
                [mediamtx_exe_path, mediamtx_config_file], 
                cwd=os.path.dirname(mediamtx_exe_path)
            )
            print(f"✅ Started MediaMTX server with PID {mediamtx_process.pid}")
            print(f"📁 Config file: {mediamtx_config_file}")
            return True
        else:
            print("ℹ️ MediaMTX is already running")
            return True
    except Exception as e:
        print(f"❌ Error starting MediaMTX: {e}")
        return False
    
def stop_mediamtx():
    """Stop the MediaMTX server process"""
    global mediamtx_process
    try:
        if mediamtx_process is not None and mediamtx_process.poll() is None:
            print(f"🛑 Stopping MediaMTX server (PID: {mediamtx_process.pid})")
            mediamtx_process.terminate()
            
            # Wait for graceful shutdown
            try:
                mediamtx_process.wait(timeout=5)
                print("✅ MediaMTX server stopped gracefully")
            except subprocess.TimeoutExpired:
                print("⚠️ Force killing MediaMTX server")
                mediamtx_process.kill()
                mediamtx_process.wait()
                print("✅ MediaMTX server force stopped")
            
            mediamtx_process = None
            return True
        else:
            print("ℹ️ MediaMTX is not running")
            return True
    except Exception as e:
        print(f"❌ Error stopping MediaMTX: {e}")
        return False

def restart_mediamtx():
    """Restart the MediaMTX server process"""
    print("🔄 Restarting MediaMTX server...")
    stop_mediamtx()
    time.sleep(2)  # Longer pause before restart to ensure clean shutdown
    result = start_mediamtx()
    if result:
        print("✅ MediaMTX server restarted successfully")
    else:
        print("❌ Failed to restart MediaMTX server")
    return result

def get_mediamtx_status():
    """Get MediaMTX server status"""
    global mediamtx_process
    if mediamtx_process is None:
        return {"running": False, "pid": None, "status": "not_started"}
    elif mediamtx_process.poll() is None:
        return {"running": True, "pid": mediamtx_process.pid, "status": "running"}
    else:
        return {"running": False, "pid": None, "status": "stopped"}

# FRP Configuration Manager Class
class FRPConfigManager:
    def __init__(self):
        self.config = {
            'common': {
                'server_addr': '103.189.89.135',
                'server_port': 7000,
                'log_level': 'info',
                'log_file': 'frpc.log',
                'log_max_days': 3
            },
            'tcp_webrtc': {
                'type': 'tcp',
                'remote_port': 10556,
                'local_ip': '127.0.0.1',
                'local_port': 8189
            },
            'udp_webrtc': {
                'type': 'udp',
                'remote_port': 10555,
                'local_ip': '127.0.0.1',
                'local_port': 8889
            },
            'ice_webrtc': {
                'type': 'tcp',
                'remote_port': 10556,
                'local_ip': '127.0.0.1',
                'local_port': 8189
            },
            'hls': {
                'type': 'tcp',
                'remote_port': 10557,
                'local_ip': '127.0.0.1',
                'local_port': 8888
            }
        }
    
    def get_config(self):
        return self.config
    
    def update_section(self, section_name, section_data):
        if section_name in self.config:
            self.config[section_name].update(section_data)
    
    def validate_config(self):
        errors = []
        # Basic validation
        if not self.config['common'].get('server_addr'):
            errors.append('Server address is required')
        if not self.config['common'].get('server_port'):
            errors.append('Server port is required')
        
        return {
            'valid': len(errors) == 0,
            'errors': errors
        }
    
    def generate_frp_ini(self, file_path='frpc.ini'):
        try:
            with open(file_path, 'w') as f:
                f.write('[common]\n')
                for key, value in self.config['common'].items():
                    f.write(f'{key} = {value}\n')
                f.write('\n')
                
                # Add TCP WebRTC
                f.write('[tcp_webrtc]\n')
                for key, value in self.config['tcp_webrtc'].items():
                    f.write(f'{key} = {value}\n')
                f.write('\n')
                
                # Add UDP WebRTC
                f.write('[udp_webrtc]\n')
                for key, value in self.config['udp_webrtc'].items():
                    f.write(f'{key} = {value}\n')
                f.write('\n')
                
                # Add ICE WebRTC
                f.write('[ice_webrtc]\n')
                for key, value in self.config['ice_webrtc'].items():
                    f.write(f'{key} = {value}\n')
                f.write('\n')
                
                # Add HLS
                f.write('[hls]\n')
                for key, value in self.config['hls'].items():
                    f.write(f'{key} = {value}\n')
            
            return True
        except Exception as e:
            print(f"Error generating FRP INI: {e}")
            return False
    
    def reset_to_defaults(self):
        self.__init__()
        return True
    
    def get_config_summary(self):
        return {
            'server': f"{self.config['common']['server_addr']}:{self.config['common']['server_port']}",
            'tunnels': len([k for k in self.config.keys() if k != 'common']),
            'log_level': self.config['common']['log_level']
        }

# Initialize FRP config manager
frp_config_manager = FRPConfigManager()

# FRP Configuration API Routes
@app.route('/frp-config')
def frp_config_page():
    """Serve the FRP configuration page"""
    return send_file('frp_config.html')

@app.route('/api/frp/config', methods=['GET'])
def get_frp_config():
    """Get current FRP configuration"""
    try:
        config = frp_config_manager.get_config()
        return jsonify({
            'success': True,
            'config': config
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/api/frp/config', methods=['POST'])
def update_frp_config():
    """Update FRP configuration"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'message': 'No configuration data provided'
            }), 400
        
        # Update each section
        for section_name, section_data in data.items():
            if section_name in frp_config_manager.config:
                frp_config_manager.update_section(section_name, section_data)
        
        # Validate configuration
        validation = frp_config_manager.validate_config()
        if not validation['valid']:
            return jsonify({
                'success': False,
                'message': 'Configuration validation failed',
                'errors': validation['errors']
            }), 400
        
        return jsonify({
            'success': True,
            'message': 'Configuration updated successfully'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/api/frp/generate-ini', methods=['POST'])
def generate_frp_ini():
    """Generate FRP INI configuration file"""
    try:
        success = frp_config_manager.generate_frp_ini()
        if success:
            # Read the generated INI file
            with open('frpc.ini', 'r') as f:
                ini_content = f.read()
            
            return jsonify({
                'success': True,
                'message': 'INI file generated successfully',
                'ini_content': ini_content
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to generate INI file'
            }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/api/frp/reset', methods=['POST'])
def reset_frp_config():
    """Reset FRP configuration to defaults"""
    try:
        success = frp_config_manager.reset_to_defaults()
        if success:
            return jsonify({
                'success': True,
                'message': 'Configuration reset to defaults'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to reset configuration'
            }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/api/frp/validate', methods=['GET'])
def validate_frp_config():
    """Validate current FRP configuration"""
    try:
        validation = frp_config_manager.validate_config()
        return jsonify({
            'success': True,
            'validation': validation
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/api/frp/summary', methods=['GET'])
def get_frp_summary():
    """Get FRP configuration summary"""
    try:
        summary = frp_config_manager.get_config_summary()
        return jsonify({
            'success': True,
            'summary': summary
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/api/frp/store-ini', methods=['POST'])
def store_frp_ini():
    """Store FRP INI file in frp-c subfolder"""
    try:
        # Create frp-c directory if it doesn't exist
        frp_c_dir = os.path.join(os.getcwd(), 'frp-c')
        if not os.path.exists(frp_c_dir):
            os.makedirs(frp_c_dir)
        
        # Generate INI content
        success = frp_config_manager.generate_frp_ini('frp-c/frpc.ini')
        
        if success:
            # Verify the file was created
            ini_path = os.path.join(frp_c_dir, 'frpc.ini')
            if os.path.exists(ini_path):
                file_size = os.path.getsize(ini_path)
                return jsonify({
                    'success': True,
                    'message': f'FRP INI file stored successfully in frp-c/ folder',
                    'file_path': ini_path,
                    'file_size': file_size
                })
            else:
                return jsonify({
                    'success': False,
                    'message': 'INI file generation succeeded but file not found'
                }), 500
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to generate INI file'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error storing INI file: {str(e)}'
        }), 500

def restart_frp_client():
    """Restart FRP client service"""
    try:
        # This would implement actual FRP client restart logic
        # For now, just return success
        return True
    except Exception as e:
        print(f"Error restarting FRP client: {e}")
        return False

def get_frp_status():
    """Get FRP client status"""
    try:
        # This would implement actual FRP status checking
        # For now, return a mock status
        return {
            'running': False,
            'pid': None,
            'status': 'not_started'
        }
    except Exception as e:
        print(f"Error getting FRP status: {e}")
        return {
            'running': False,
            'pid': None,
            'status': 'error'
        }

def start_frp_client():
    """Start FRP client service"""
    try:
        # This would implement actual FRP client start logic
        return True
    except Exception as e:
        print(f"Error starting FRP client: {e}")
        return False

def stop_frp_client():
    """Stop FRP client service"""
    try:
        # This would implement actual FRP client stop logic
        return True
    except Exception as e:
        print(f"Error stopping FRP client: {e}")
        return False

@app.route('/api/frp/restart', methods=['POST'])
def restart_frp_client_api():
    """Restart FRP client service"""
    try:
        success = restart_frp_client()
        if success:
            return jsonify({
                'success': True,
                'message': 'FRP client restarted successfully'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to restart FRP client'
            }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error restarting FRP client: {str(e)}'
        }), 500

@app.route('/api/frp/status', methods=['GET'])
def get_frp_status_api():
    """Get FRP client status"""
    try:
        status = get_frp_status()
        return jsonify({
            'success': True,
            'status': status
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error getting FRP status: {str(e)}'
        }), 500

@app.route('/api/frp/start', methods=['POST'])
def start_frp_client_api():
    """Start FRP client service"""
    try:
        success = start_frp_client()
        if success:
            return jsonify({
                'success': True,
                'message': 'FRP client started successfully'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to start FRP client'
            }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error starting FRP client: {str(e)}'
        }), 500

@app.route('/api/frp/stop', methods=['POST'])
def stop_frp_client_api():
    """Stop FRP client service"""
    try:
        success = stop_frp_client()
        if success:
            return jsonify({
                'success': True,
                'message': 'FRP client stopped successfully'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to stop FRP client'
            }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error stopping FRP client: {str(e)}'
        }), 500


@app.route('/api/update-config-v2', methods=['POST'])  # Changed route
def update_config_v2():  # Changed function name
    try:
        config_data = request.get_json()
        config_file_path = "config.json"
        
        # Overwrite the config file
        with open(config_file_path, "w") as f:
            json.dump(config_data, f, indent=2)
        
        return jsonify({
            "status": "success", 
            "message": f"Config updated with {len(config_data.get('cameras', []))} cameras"
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def initialize_video_monitoring():
    """Initialize video file monitoring system"""
    global file_monitor
    
    # Load video status on startup
    load_video_status()
    
    # Process any pending videos from previous runs
    process_pending_videos()
    
    # Determine video directory to monitor
    video_base_dir = base_video_path if base_video_path else "videos"
    
    if not os.path.exists(video_base_dir):
        os.makedirs(video_base_dir)
        print(f"📁 Created video directory: {video_base_dir}")
    
    # Start file monitor (disabled for now - using cleanup thread only)
    # file_monitor = VideoFileMonitor(video_base_dir, video_upload_callback)
    # file_monitor.start()
    
    print(f"🎬 Video monitoring system initialized for: {video_base_dir} (using cleanup thread only)")

def main():
    """Main function"""
    # Load cameras first
    load_cameras()
    
    # Start MediaMTX server
    print("🚀 Starting MediaMTX server...")
    if start_mediamtx():
        print("✅ MediaMTX server started successfully")
    else:
        print("❌ Failed to start MediaMTX server")
    
    # Start the dedicated upload thread
    start_dedicated_upload_thread()
    
    # Initialize video monitoring system
    initialize_video_monitoring()
    
    # Start cleanup thread for file management and uploads
    start_cleanup_thread()
    
    # Start the Flask server
    print("Starting Camera Management System...")
    print(f"Web interface available at: http://{server_config['host']}:{server_config['port']}")
    print("Press Ctrl+C to stop the server")
    
    try:
        app.run(host=server_config['host'], port=server_config['port'], debug=server_config['debug'])
    except KeyboardInterrupt:
        print("\nShutting down...")
        # Stop MediaMTX server
        print("🛑 Stopping MediaMTX server...")
        stop_mediamtx()
        # Stop dedicated upload thread
        stop_dedicated_upload_thread()
        
        # Stop file monitor
        if file_monitor:
            print("🛑 Stopping video file monitor...")
            file_monitor.stop()
        
        # Stop cleanup thread
        print("🛑 Stopping cleanup thread...")
        stop_cleanup_thread()
        
        # Stop all active threads
        for camera_id in list(active_threads.keys()):
            stop_camera_thread(camera_id)
        print("All threads stopped.")

if __name__ == "__main__":
    main() 